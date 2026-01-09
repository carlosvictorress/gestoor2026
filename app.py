# FORÇANDO A ATUALIZAÇÃO DO CACHE
# ===================================================================
# PARTE 1: Importações de Bibliotecas
# ===================================================================
import os
import io
from dotenv import load_dotenv
load_dotenv()
import csv
import uuid
import locale
import qrcode
import base64
import re
import math
import face_recognition
import json
from sqlalchemy import or_

from reportlab.platypus import KeepTogether

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from flask import make_response
from reportlab.lib.pagesizes import A4, landscape

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer # (Você já deve ter)
from reportlab.lib.styles import getSampleStyleSheet # (Você já deve ter)
from reportlab.lib import colors # (Você já deve ter)
from reportlab.lib.units import cm # (Você já deve ter)
from utils import cabecalho_e_rodape # (Você já deve ter)
from utils import currency_filter_br, upload_arquivo_para_nuvem
from utils import identificar_servidor_por_rosto, haversine

from flask_mail import Message # Adicione esta
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadTimeSignature # Adicione esta

from functools import wraps
from datetime import datetime, timedelta
from flask import session, flash, redirect, url_for
from flask_talisman import Talisman

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    make_response,
    send_from_directory,
    Response,
    abort,
    jsonify,
)

import click

from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_bcrypt import Bcrypt
from sqlalchemy import func, or_, and_
from werkzeug.utils import secure_filename
from num2words import num2words
from utils import gerar_encoding_facial, comparar_rostos, limpar_cpf

# Importações para o gerador de PDF (ReportLab)
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
    BaseDocTemplate,
    Frame,
    PageTemplate,
    KeepTogether,
    ListFlowable,
    HRFlowable,
)
from reportlab.pdfgen import canvas as canvas_lib
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.lib.units import cm, inch

# Configura o locale para o português do Brasil
try:
    locale.setlocale(locale.LC_TIME, "pt_BR.UTF-8")
except locale.Error:
    print("Locale pt_BR.UTF-8 não encontrado, usando o padrão do sistema.")

# ===================================================================
# PARTE 2: Configuração da Aplicação e Inicialização das Extensões
# ===================================================================


app = Flask(__name__)
basedir = os.path.abspath(os.path.dirname(__file__))

app.jinja_env.filters['currency_br'] = currency_filter_br



# --- CÓDIGO CORRIGIDO ---
database_url = os.environ.get("DATABASE_URL")

# A condição foi ajustada para 'postgresql://'
if database_url and database_url.startswith("postgresql://"):
    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
else:
    # Se não houver DATABASE_URL, usa o banco de dados SQLite local
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(basedir, "servidores.db")

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = "uma-chave-secreta-muito-dificil-de-adivinhar"
app.config["UPLOAD_FOLDER"] = "uploads"

#app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.googlemail.com') # Servidor SMTP
#app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587)) # Porta (587 para TLS)
#app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'true').lower() in ['true', 'on', '1'] # Usar TLS? (True para Gmail)
#app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME') # Seu endereço de e-mail completo
#app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD') # Sua senha (ou senha de app para Gmail)
#app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_USERNAME') # O remetente padrão será seu e-mail

app.config['MAIL_SERVER'] = 'smtp.googlemail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'carlosvictor.pessoal@gmail.com'
app.config['MAIL_PASSWORD'] = 'cphvgocxclcmzxqc'  # Sua Senha de App (sem espaços)
app.config['MAIL_DEFAULT_SENDER'] = 'carlosvictor.pessoal@gmail.com'


#RAIO_PERMITIDO_METROS = 100
app.config['RAIO_PERMITIDO_METROS'] = 100

s = URLSafeTimedSerializer(app.config['SECRET_KEY'])

# --- Inicialização das Extensões ---
from extensions import db, bcrypt

from flask_mail import Mail
mail = Mail(app)

db.init_app(app)
bcrypt.init_app(app)
migrate = Migrate(app, db)

# ===================================================================
# PARTE 3: Importação dos Modelos
# ===================================================================
from models import *




# ===================================================================
# PARTE 4: Comandos CLI, Funções Globais e Decoradores
# ===================================================================


def role_required(*roles):
    """
    Decorador que verifica se o usuário logado tem uma das permissões necessárias.
    O 'admin' sempre tem acesso.
    """

    def wrapper(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # O admin sempre tem acesso a tudo
            if "admin" in roles or session.get("role") == "admin":
                return f(*args, **kwargs)

            # Verifica se o papel do usuário está na lista de papéis permitidos
            if session.get("role") not in roles:
                flash("Você não tem permissão para acessar esta página.", "danger")
                return redirect(
                    url_for("dashboard")
                )  # Redireciona para o painel principal
            return f(*args, **kwargs)

        return decorated_function

    return wrapper


def check_license(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Defina aqui as rotas que podem ser acessadas mesmo com a licença expirada
        allowed_routes = [
            "login",
            "logout",
            "renovar_licenca",
            "admin_licenca",
            "static",
            "uploaded_file",
        ]

        if request.endpoint in allowed_routes:
            return f(*args, **kwargs)

        # Busca a licença no banco de dados
        licenca = License.query.first()

        # Se a licença não existe ou está expirada
        if not licenca or licenca.expiration_date < datetime.utcnow():
            # Permite que o admin acesse para poder renovar
            if session.get("role") == "admin":
                return f(*args, **kwargs)

            # Para outros usuários, exibe a mensagem e redireciona
            flash(
                "Sua licença de uso do sistema expirou. Por favor, renove sua assinatura para continuar.",
                "warning",
            )
            return redirect(url_for("renovar_licenca"))

        # Se a licença estiver válida, permite o acesso
        return f(*args, **kwargs)

    return decorated_function


@app.cli.command("init-db")
def init_db_command():
    db.create_all()
    os.makedirs(os.path.join(app.config["UPLOAD_FOLDER"], "documentos"), exist_ok=True)
    print("Banco de dados e pastas de uploads inicializados.")


@app.cli.command("create-admin")
def create_admin_command():
    with app.app_context():
        username = input("Digite o nome de usuário para o admin: ")
        password = input("Digite a senha para o admin: ")

        # --- NOVO CAMPO ADICIONADO ---
        email = input("Digite o E-MAIL para o admin: ")
        # ---------------------------

        # Lista as secretarias disponíveis
        secretarias = Secretaria.query.all()
        if not secretarias:
            print("Erro: Nenhuma secretaria encontrada. Crie uma primeiro com 'flask add-secretaria'.")
            return

        print("Selecione a secretaria para este usuário:")
        for i, sec in enumerate(secretarias):
            print(f"  {i + 1}: {sec.nome}")

        while True:
            try:
                choice = int(input("Digite o número da secretaria: "))
                if 1 <= choice <= len(secretarias):
                    secretaria_selecionada = secretarias[choice - 1]
                    break
                else:
                    print("Seleção inválida. Tente novamente.")
            except ValueError:
                print("Por favor, digite um número.")

        # Verifica se o usuário já existe
        user = User.query.filter_by(username=username).first()
        if user:
            print(f"Usuário '{username}' já existe. Atualizando secretaria e e-mail.")
            user.secretaria_id = secretaria_selecionada.id
            user.email = email # <-- ATUALIZA O EMAIL
        else:
            hashed_password = bcrypt.generate_password_hash(password).decode("utf-8")
            new_user = User(
                username=username, 
                password_hash=hashed_password, 
                role="admin",
                secretaria_id=secretaria_selecionada.id,
                email=email # <-- CAMPO ADICIONADO AQUI
            )
            db.session.add(new_user)

        db.session.commit()
        print(f"Usuário administrador '{username}' criado/atualizado com sucesso na secretaria '{secretaria_selecionada.nome}'!")


@app.cli.command("add-secretaria")
@click.argument("nome")
def add_secretaria_command(nome):
    """Cria uma nova secretaria a partir da linha de comando."""
    if Secretaria.query.filter_by(nome=nome).first():
        print(f"Erro: A secretaria '{nome}' já existe.")
        return
    nova_secretaria = Secretaria(nome=nome)
    db.session.add(nova_secretaria)
    db.session.commit()
    print(f"Secretaria '{nome}' criada com sucesso!")






@app.cli.command("init-licence")
def init_licence_command():
    """Cria uma licença inicial para o sistema."""
    with app.app_context():
        if License.query.first():
            print("Erro: A licença já existe na base de dados.")
            return

        # Cria uma licença que expira em 30 dias a partir de agora
        expiration = datetime.utcnow() + timedelta(days=30)
        new_licence = License(expiration_date=expiration)
        db.session.add(new_licence)
        db.session.commit()
        print(f"Licença inicial criada com sucesso! Expira em: {expiration.strftime('%d/%m/%Y')}")





@app.context_processor
def inject_year():
    return {"current_year": datetime.utcnow().year}

@app.template_filter('today_date')
def today_date_filter(s):
    """Filtro Jinja para retornar a data de hoje no formato YYYY-MM-DD."""
    # 'datetime' já deve estar importado no topo do seu app.py
    return datetime.now().strftime('%Y-%m-%d')


def registrar_log(action):
    try:
        if "logged_in" in session:
            username = session.get("username", "Anônimo")
            ip_address = request.remote_addr
            log_entry = Log(username=username, action=action, ip_address=ip_address)
            db.session.add(log_entry)
            db.session.commit()
    except Exception as e:
        print(f"Erro ao registrar log: {e}")
        db.session.rollback()





def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "logged_in" not in session:
            flash("Por favor, faça login para acessar esta página.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Você não tem permissão para acessar esta página.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated_function


def fleet_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") not in ["admin", "fleet"]:
            flash("Você não tem permissão para acessar este módulo.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated_function


def cabecalho_e_rodape(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 9)
    canvas.drawString(
        2 * cm, 1.5 * cm, f"Emitido em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    )
    canvas.drawRightString(doc.width + doc.leftMargin, 1.5 * cm, f"Página {doc.page}")

    if doc.page == 1:
        image_path = os.path.join(basedir, "static", "timbre.jpg")
        if os.path.exists(image_path):
            canvas.drawImage(
                image_path,
                2 * cm,
                A4[1] - 2.5 * cm,  # A4[1] é a altura da página
                width=17 * cm,
                height=2.2 * cm,
                preserveAspectRatio=True,
                mask="auto",
            )

    canvas.restoreState()


# ===================================================================
# PARTE 5: Definição das Rotas da Aplicação
# ===================================================================

@app.route("/documento/download/<int:documento_id>")
@login_required
@role_required("RH", "admin")
def download_documento(documento_id):
    # Nota: Assumimos que o modelo 'Documento' está definido em 'models.py'
    doc = Documento.query.get_or_404(documento_id)
    # Define o caminho para a pasta 'uploads/documentos'
    upload_path = os.path.join(current_app.config["UPLOAD_FOLDER"], "documentos")

    try:
        from flask import current_app # Garante que a variável esteja disponível

        # Usa send_from_directory para servir o arquivo
        return send_from_directory(
            upload_path, 
            doc.filename, 
            as_attachment=True, 
            download_name=doc.description
        )
    except FileNotFoundError:
        flash("Erro: Arquivo não encontrado no servidor.", "danger")
        # Redireciona de volta para a edição do servidor
        return redirect(url_for("editar_servidor", id=doc.servidor_id))

@app.route("/documento/delete/<int:documento_id>")
@login_required
@admin_required
@role_required("RH", "admin")
def delete_documento(documento_id):
    doc = Documento.query.get_or_404(documento_id)
    servidor_id = doc.servidor_id # Pega o ID do servidor para o redirecionamento

    try:
        # 1. Remove o arquivo físico
        doc_path = os.path.join(
            current_app.config["UPLOAD_FOLDER"], "documentos", doc.filename
        )
        if os.path.exists(doc_path):
            os.remove(doc_path)

        # 2. Remove o registro do banco de dados
        db.session.delete(doc)
        db.session.commit()
        registrar_log(f'Excluiu o documento "{doc.description}" do servidor {servidor_id}.')
        flash("Documento excluído com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir o documento: {e}", "danger")

    return redirect(url_for("editar_servidor", id=servidor_id))


@app.route("/reset_password", methods=["GET", "POST"])
def reset_request():
    """ Rota para solicitar a redefinição de senha via e-mail. """
    if 'logged_in' in session: # Se já estiver logado, não faz sentido redefinir
        return redirect(url_for('dashboard'))

    if request.method == "POST":
        email = request.form.get("email")
        # Busca o usuário pelo e-mail fornecido
        user = User.query.filter_by(email=email).first()

        if user:
            # Se o usuário existir, tenta gerar e enviar o e-mail
            try:
                # Gera um token seguro contendo o ID do usuário.
                # O 'salt' adiciona uma camada extra de segurança.
                # O token será válido por 3600 segundos (1 hora).
                token = s.dumps(user.id, salt='password-reset-salt')

                # Cria o URL completo que será enviado no e-mail
                # _external=True garante que o URL inclua http://dominio...
                reset_url = url_for('reset_token', token=token, _external=True)

                # Cria o objeto da mensagem de e-mail
                msg = Message(
                    'Redefinição de Senha - Gestoor360', # Assunto
                    recipients=[user.email] # Destinatário
                    # O remetente será o MAIL_DEFAULT_SENDER configurado
                )
                # Define o corpo do e-mail (usando f-string e triple quotes)
                msg.body = f"""Olá {user.username},

Recebemos uma solicitação para redefinir a sua senha no sistema Gestoor360.

Para criar uma nova senha, por favor, clique no link abaixo. Este link expirará em 1 hora:
{reset_url}

Se você não solicitou esta alteração, pode ignorar este e-mail com segurança. Sua senha não será alterada.

Atenciosamente,
Equipe Gestoor360
""" # <-- Fim das triple quotes para msg.body

                # Envia o e-mail usando a extensão Flask-Mail
                mail.send(msg)

                flash('Um e-mail foi enviado com as instruções para redefinir a sua senha.', 'info')
            except Exception as e:
                # Em caso de erro no envio (configuração errada, etc.), loga e avisa o usuário
                print(f"Erro ao enviar e-mail de redefinição para {email}: {e}")
                flash('Ocorreu um erro ao tentar enviar o e-mail. Por favor, tente novamente mais tarde ou contacte o suporte.', 'danger')
        else:
            # Se o e-mail não foi encontrado, mostra uma mensagem genérica por segurança
            # (para não confirmar se um e-mail está ou não cadastrado)
            flash('Se o endereço de e-mail estiver registado no nosso sistema, você receberá um link para redefinir a senha.', 'info')

        # Redireciona de volta para a página de login após a tentativa
        return redirect(url_for('login'))

    # Se o método for GET, simplesmente exibe o formulário para pedir a redefinição
    return render_template('reset_request.html')


@app.route("/reset_password/<token>", methods=["GET", "POST"])
def reset_token(token):
    """ Rota que recebe o token do e-mail e permite definir a nova senha. """
    if 'logged_in' in session: # Se já estiver logado, não faz sentido estar aqui
        return redirect(url_for('dashboard'))

    try:
        # Tenta decodificar o token, validando o 'salt' e o tempo de expiração (max_age)
        user_id = s.loads(token, salt='password-reset-salt', max_age=3600)
    except SignatureExpired:
        # Se o token expirou (passou mais de 1 hora)
        flash('O link para redefinição de senha expirou. Por favor, solicite um novo.', 'warning')
        return redirect(url_for('reset_request'))
    except (BadTimeSignature, Exception) as e:
        # Se o token é inválido (malformado, salt errado, etc.)
        print(f"Erro ao decodificar token: {e}")
        flash('O link para redefinição de senha é inválido ou já foi utilizado.', 'danger')
        return redirect(url_for('reset_request'))

    # Se o token foi válido, busca o usuário correspondente no banco
    user = User.query.get(user_id)
    if not user:
        # Se o usuário associado ao token não existe mais (pouco provável)
        flash('Usuário associado a este link não encontrado.', 'danger')
        return redirect(url_for('login'))

    # Se o formulário de nova senha for enviado (método POST)
    if request.method == "POST":
        password = request.form.get("password")
        password_confirm = request.form.get("password_confirm")

        # Verifica se as senhas foram preenchidas e se são iguais
        if not password or not password_confirm:
            flash('Por favor, preencha ambos os campos de senha.', 'warning')
            return render_template('reset_password.html', token=token)
        if password != password_confirm:
            flash('As senhas digitadas não conferem. Tente novamente.', 'warning')
            return render_template('reset_password.html', token=token)

        # Se as senhas são válidas
        try:
            # Gera o hash da nova senha
            hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
            # Atualiza o hash da senha do usuário no banco de dados
            user.password_hash = hashed_password
            db.session.commit() # Salva a alteração

            # Registra o log da alteração
            if 'registrar_log' in globals(): # Verifica se a função existe para evitar erros
              registrar_log(f'Usuário "{user.username}" redefiniu a senha via e-mail.')

            flash('Sua senha foi atualizada com sucesso! Você já pode fazer login com a nova senha.', 'success')
            return redirect(url_for('login')) # Redireciona para a página de login

        except Exception as e:
            # Em caso de erro ao salvar no banco
            db.session.rollback()
            print(f"Erro ao atualizar senha para usuário {user_id}: {e}")
            flash('Ocorreu um erro ao tentar atualizar sua senha. Por favor, tente novamente.', 'danger')
            return render_template('reset_password.html', token=token)

    # Se o método for GET, simplesmente exibe o formulário para digitar a nova senha
    return render_template('reset_password.html', token=token)



@app.route("/requerimentos/editar/<int:req_id>", methods=['GET', 'POST'])
@login_required
@role_required('RH', 'admin')
def editar_requerimento(req_id):
    requerimento = Requerimento.query.get_or_404(req_id)
    if request.method == 'POST':
        try:
            # Atualiza os dados do requerimento a partir do formulário
            requerimento.autoridade_dirigida = request.form.get('autoridade_dirigida')
            requerimento.natureza = request.form.get('natureza')
            requerimento.natureza_outro = request.form.get('natureza_outro') if requerimento.natureza == 'Outro' else None
            data_admissao_str = request.form.get('data_admissao')
            requerimento.data_admissao = datetime.strptime(data_admissao_str, '%Y-%m-%d').date() if data_admissao_str else None
            data_inicio_req_str = request.form.get('data_inicio_requerimento')
            requerimento.data_inicio_requerimento = datetime.strptime(data_inicio_req_str, '%Y-%m-%d').date() if data_inicio_req_str else None
            requerimento.duracao = request.form.get('duracao')
            requerimento.periodo_aquisitivo = request.form.get('periodo_aquisitivo')
            requerimento.informacoes_complementares = request.form.get('informacoes_complementares')
            requerimento.parecer_juridico = request.form.get('parecer_juridico')

            db.session.commit()
            flash('Requerimento atualizado com sucesso!', 'success')
            return redirect(url_for('listar_requerimentos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar o requerimento: {e}', 'danger')

    return render_template('editar_requerimento.html', requerimento=requerimento)






@app.route("/requerimentos/mudar-status-modal", methods=["POST"])
@login_required
@role_required("RH", "admin")
def mudar_status_requerimento_modal():
    try:
        req_id = request.form.get("req_id", type=int)
        novo_status = request.form.get("novo_status")

        if not req_id or not novo_status:
            flash("Dados inválidos para alterar o status.", "danger")
            return redirect(url_for("listar_requerimentos"))

        req = Requerimento.query.get_or_404(req_id)
        status_permitidos = ["Aprovado", "Recusado", "Concluído", "Em Análise"]
        if novo_status not in status_permitidos:
            flash("Status inválido.", "danger")
            return redirect(url_for("listar_requerimentos"))

        req.status = novo_status
        if novo_status in ["Aprovado", "Recusado", "Concluído"]:
            req.data_conclusao = datetime.now().date()
        else:
            req.data_conclusao = None

        db.session.commit()
        registrar_log(
            f'Alterou o status do requerimento ID {req.id} para "{novo_status}".'
        )
        flash(
            f'Status do requerimento #{req.id} alterado para "{novo_status}" com sucesso!',
            "info",
        )

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao alterar status: {e}", "danger")

    return redirect(url_for("listar_requerimentos"))


@app.route("/usuarios")
@login_required
@admin_required
def lista_usuarios():
    usuarios = User.query.order_by(User.username).all()
    # Busca a nova lista de secretarias
    secretarias = Secretaria.query.order_by(Secretaria.nome).all()
    # Envia as duas listas para o template
    return render_template("usuarios.html", usuarios=usuarios, secretarias=secretarias)


@app.route("/secretarias/add", methods=["POST"])
@login_required
@admin_required
def add_secretaria():
    nome_secretaria = request.form.get("nome")
    if not nome_secretaria:
        flash("O nome da secretaria é obrigatório.", "warning")
        return redirect(url_for("lista_usuarios"))

    # Verifica se a secretaria já existe
    existente = Secretaria.query.filter_by(nome=nome_secretaria).first()
    if existente:
        flash("Essa secretaria já está cadastrada.", "danger")
        return redirect(url_for("lista_usuarios"))

    nova_secretaria = Secretaria(nome=nome_secretaria)
    db.session.add(nova_secretaria)
    db.session.commit()
    registrar_log(f'Cadastrou a secretaria: "{nome_secretaria}".')
    flash("Secretaria cadastrada com sucesso!", "success")
    return redirect(url_for("lista_usuarios"))


@app.route("/secretarias/delete/<int:id>")
@login_required
@admin_required
def delete_secretaria(id):
    secretaria_para_excluir = Secretaria.query.get_or_404(id)
    try:
        nome_sec = secretaria_para_excluir.nome
        db.session.delete(secretaria_para_excluir)
        db.session.commit()
        registrar_log(f'Excluiu a secretaria: "{nome_sec}".')
        flash(f'Secretaria "{nome_sec}" excluída com sucesso.', "success")
    except Exception as e:
        db.session.rollback()
        # Este erro pode acontecer se um usuário estiver vinculado a esta secretaria no futuro
        flash(
            f"Não foi possível excluir a secretaria. Verifique se não há vínculos. Erro: {e}",
            "danger",
        )

    return redirect(url_for("lista_usuarios"))


@app.route("/usuarios/add", methods=["POST"])
@login_required
@admin_required
def add_usuario():
    username = request.form.get("username")
    password = request.form.get("password")
    # --- LINHA ADICIONADA ---
    email = request.form.get("email")
    # ---------------------------
    role = request.form.get("role", "operador")
    secretaria_id = request.form.get("secretaria_id", type=int)

    # --- VALIDAÇÃO ATUALIZADA ---
    if not all([username, password, email, secretaria_id]):
        flash(
            "Todos os campos (Usuário, E-mail, Senha e Secretaria) são obrigatórios.", "warning"
        )
        return redirect(url_for("lista_usuarios"))
    # -----------------------------

    user_exists = User.query.filter_by(username=username).first()
    if user_exists:
        flash("Este nome de usuário já existe.", "danger")
        return redirect(url_for("lista_usuarios"))
        
    # --- VERIFICAÇÃO DE E-MAIL ADICIONADA ---
    email_exists = User.query.filter_by(email=email).first()
    if email_exists:
        flash("Este e-mail já está em uso.", "danger")
        return redirect(url_for("lista_usuarios"))
    # -------------------------------------

    hashed_password = bcrypt.generate_password_hash(password).decode("utf-8")
    new_user = User(
        username=username,
        # --- E-MAIL ADICIONADO ---
        email=email,
        # --------------------------
        password_hash=hashed_password,
        role=role,
        secretaria_id=secretaria_id,
    )

    db.session.add(new_user)
    db.session.commit()
    registrar_log(f'Criou o usuário: "{username}".')
    flash(f'Usuário "{username}" criado com sucesso!', "success")
    return redirect(url_for("lista_usuarios"))


@app.route("/usuarios/editar/<int:id>", methods=["GET", "POST"])
@login_required
@admin_required
def editar_usuario(id):
    user = User.query.get_or_404(id)
    # Pega a lista de secretarias para o dropdown
    secretarias = Secretaria.query.order_by(Secretaria.nome).all()

    if request.method == "POST":
        # --- LÓGICA DE VERIFICAÇÃO DO ÚLTIMO ADMIN ---
        if (
            user.role == "admin"
            and User.query.filter_by(role="admin").count() == 1
            and request.form.get("role") != "admin" # Se a tentativa for mudar o role
        ):
            flash(
                "Não é possível remover o status de administrador do último admin do sistema.",
                "danger",
            )
            # Recarrega a página com os dados atuais
            return render_template("editar_usuario.html", user=user, secretarias=secretarias)
        
        # --- VERIFICAÇÃO DE USERNAME E E-MAIL ---
        new_username = request.form.get("username")
        new_email = request.form.get("email")

        # Verifica se o username já está em uso por OUTRO usuário
        user_exists = User.query.filter(
            User.username == new_username, User.id != id
        ).first()
        if user_exists:
            flash("Este nome de usuário já está em uso.", "danger")
            return render_template("editar_usuario.html", user=user, secretarias=secretarias)

        # Verifica se o E-MAIL já está em uso por OUTRO usuário
        email_exists = User.query.filter(
            User.email == new_email, User.id != id
        ).first()
        if email_exists:
            flash("Este e-mail já está em uso por outro usuário.", "danger")
            return render_template("editar_usuario.html", user=user, secretarias=secretarias)
            
        # --- ATUALIZA OS DADOS DO USUÁRIO ---
        try:
            user.username = new_username
            user.email = new_email  # <- CAMPO ADICIONADO
            user.role = request.form.get("role")
            user.secretaria_id = request.form.get("secretaria_id", type=int) # <- CAMPO ADICIONADO
            
            new_password = request.form.get("password")
            if new_password:
                user.password_hash = bcrypt.generate_password_hash(new_password).decode(
                    "utf-8"
                )
                
            db.session.commit()
            registrar_log(f'Editou o usuário: "{user.username}".')
            flash("Usuário atualizado com sucesso!", "success")
            return redirect(url_for("lista_usuarios"))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao atualizar usuário: {e}", "danger")

    # (GET) Apenas renderiza a página de edição
    return render_template("editar_usuario.html", user=user, secretarias=secretarias)


@app.route("/logs")
@login_required
@admin_required
def ver_logs():
    page = request.args.get("page", 1, type=int)
    logs_pagination = Log.query.order_by(Log.timestamp.desc()).paginate(
        page=page, per_page=25, error_out=False
    )
    return render_template("logs.html", logs=logs_pagination)


@app.route("/admin/licenca", methods=["GET", "POST"])
@login_required
@admin_required
def admin_licenca():
    licenca = License.query.first_or_404()
    if request.method == "POST":
        nova_chave = str(uuid.uuid4())
        licenca.renewal_key = nova_chave
        db.session.commit()
        registrar_log("Gerou uma nova chave de renovação.")
        flash("Nova chave de renovação gerada com sucesso!", "success")
        return redirect(url_for("admin_licenca"))
    return render_template("admin_licenca.html", licenca=licenca)


@app.route("/renovar", methods=["GET", "POST"])
@login_required
def renovar_licenca():
    licenca = License.query.first_or_404()
    if request.method == "POST":
        chave_inserida = request.form.get("renewal_key")
        if licenca.renewal_key and licenca.renewal_key == chave_inserida:
            licenca.expiration_date = datetime.utcnow() + timedelta(days=31)
            licenca.renewal_key = None
            db.session.commit()
            registrar_log("Renovou a licença do sistema com sucesso.")
            flash("Licença renovada com sucesso! Obrigado.", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Chave de renovação inválida ou já utilizada.", "danger")

    if licenca.expiration_date >= datetime.utcnow() and session.get("role") != "admin":
        return redirect(url_for("dashboard"))

    return render_template("renovar_licenca.html")


@app.route("/relatorio/servidores/pdf")
@login_required
def gerar_relatorio_pdf():
    servidores = Servidor.query.order_by(Servidor.nome).all()

    buffer = io.BytesIO()
    doc = BaseDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=3 * cm,
        bottomMargin=2.5 * cm,
    )

    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
    template = PageTemplate(
        id="main_template", frames=[frame], onPage=cabecalho_e_rodape
    )
    doc.addPageTemplates([template])

    styles = getSampleStyleSheet()
    p_style = ParagraphStyle(
        name="CustomNormal", parent=styles["Normal"], alignment=TA_CENTER, fontSize=8
    )
    header_style = ParagraphStyle(
        name="CustomHeader",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=9,
        fontName="Helvetica-Bold",
    )

    story = [
        Paragraph("Relatório Geral de Servidores", styles["h1"]),
        Spacer(1, 1 * cm),
    ]

    if not servidores:
        story.append(Paragraph("Nenhum servidor cadastrado.", styles["Normal"]))
    else:
        table_data = [
            [
                Paragraph(h, header_style)
                for h in ["Nome", "CPF", "Função", "Lotação", "Vínculo", "Telefone"]
            ]
        ]
        for s in servidores:
            row = [
                Paragraph(s.nome or "", p_style),
                Paragraph(s.cpf or "", p_style),
                Paragraph(s.funcao or "", p_style),
                Paragraph(s.lotacao or "", p_style),
                Paragraph(s.tipo_vinculo or "", p_style),
                Paragraph(s.telefone or "", p_style),
            ]
            table_data.append(row)

        table = Table(
            table_data, colWidths=[7 * cm, 3.5 * cm, 4 * cm, 4 * cm, 3.5 * cm, 3 * cm]
        )
        style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#004d40")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
        table.setStyle(style)
        story.append(table)

    doc.build(story)
    buffer.seek(0)

    response = make_response(buffer.getvalue())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = (
        f'inline; filename=relatorio_servidores_{datetime.now().strftime("%Y-%m-%d")}.pdf'
    )

    registrar_log("Gerou o PDF do Relatório Geral de Servidores.")
    return response


@app.route("/combustivel/relatorio/mensal/selecionar")
@login_required
def pagina_relatorio_mensal():
    return render_template("relatorio_mensal.html")


@app.route("/relatorio/combustivel/tce-pi")
@login_required
@role_required("Combustivel", "admin")
def relatorio_combustivel_tce_pi():
    """
    Gera um relatório de abastecimento no formato CSV exigido pelo TCE-PI.
    """
    try:
        abastecimentos = (
            Abastecimento.query.join(
                Motorista, Abastecimento.motorista_id == Motorista.id
            )
            .join(Veiculo)
            .order_by(Abastecimento.data.asc())
            .all()
        )

        output = io.StringIO()
        writer = csv.writer(output, delimiter=";")

        header = [
            "unidade_gestora",
            "exercicio",
            "mes_referencia",
            "numero_notafiscal",
            "data_notafiscal",
            "cpf_condutor",
            "nome_condutor",
            "placa_veiculo",
            "quilometragem",
            "tipo_combustivel",
            "quantidade_combustivel",
            "valor_unitario",
            "valor_total",
            "cnpj_fornecedor",
        ]
        writer.writerow(header)

        for r in abastecimentos:
            row = [
                "",
                r.data.year,
                r.data.month,
                "",
                r.data.strftime("%Y-%m-%d"),
                r.motorista.cpf,
                r.motorista.nome,
                r.veiculo.placa,
                f"{r.quilometragem:.1f}".replace(".", ","),
                r.tipo_combustivel,
                f"{r.litros:.2f}".replace(".", ","),
                f"{r.valor_litro:.2f}".replace(".", ","),
                f"{r.valor_total:.2f}".replace(".", ","),
                "",
            ]
            writer.writerow(row)

        csv_content = output.getvalue()
        response = Response(
            csv_content.encode("utf-8-sig"),
            mimetype="text/csv",
            headers={
                "Content-Disposition": "attachment;filename=Relatorio_Abastecimento_TCE-PI.csv"
            },
        )
        registrar_log("Gerou o relatório de abastecimento para o TCE-PI.")
        return response

    except Exception as e:
        flash(f"Ocorreu um erro ao gerar o relatório: {e}", "danger")
        return redirect(url_for("lancar_abastecimento"))


@app.route("/relatorio/veiculos/selecionar")
@login_required
@role_required("Combustivel", "admin")
def selecionar_relatorio_veiculos():
    return render_template("selecionar_relatorio_veiculos.html")


@app.route("/relatorio/veiculos/gerar", methods=["POST"])
@login_required
@fleet_required
@role_required("Combustivel", "admin")
def gerar_relatorio_veiculos_mensal():
    try:
        ano = int(request.form.get("ano"))
        mes = int(request.form.get("mes"))
        orgao_filtro = request.form.get("orgao")

        data_inicio = datetime(ano, mes, 1)
        if mes == 12:
            data_fim = datetime(ano + 1, 1, 1)
        else:
            data_fim = datetime(ano, mes + 1, 1)

        query_veiculos = Veiculo.query

        if orgao_filtro and orgao_filtro != "todos":
            query_veiculos = query_veiculos.filter(Veiculo.orgao == orgao_filtro)

        veiculos = query_veiculos.order_by(Veiculo.modelo).all()
        dados_relatorio = []

        for veiculo in veiculos:
            abastecimentos_do_mes = (
                Abastecimento.query.filter(
                    Abastecimento.veiculo_placa == veiculo.placa,
                    Abastecimento.data >= data_inicio,
                    Abastecimento.data < data_fim,
                )
                .order_by(Abastecimento.quilometragem.asc())
                .all()
            )

            if abastecimentos_do_mes:
                km_anterior_por_veiculo = {}
                ultimo_abastecimento_anterior = (
                    Abastecimento.query.filter(
                        Abastecimento.veiculo_placa == veiculo.placa,
                        Abastecimento.data < data_inicio,
                    )
                    .order_by(Abastecimento.quilometragem.desc())
                    .first()
                )

                if ultimo_abastecimento_anterior:
                    km_anterior_por_veiculo[veiculo.placa] = (
                        ultimo_abastecimento_anterior.quilometragem
                    )

                for r in abastecimentos_do_mes:
                    placa = r.veiculo.placa
                    km_inicial = km_anterior_por_veiculo.get(placa, 0)
                    km_final = r.quilometragem
                    dados_relatorio.append(
                        {
                            "modelo": veiculo.modelo,
                            "placa": placa,
                            "renavam": veiculo.renavam or "",
                            "ano_fab": veiculo.ano_fabricacao or "",
                            "ano_mod": veiculo.ano_modelo or "",
                            "tipo_veiculo": r.veiculo.tipo or "AUTOMOVEL",
                            "orgao_localizacao": veiculo.orgao or "",
                            "qtde_abastecimento": f"{r.litros:.2f}".replace(".", ","),
                            "combustivel": r.tipo_combustivel,
                            "km_inicial_mes": (
                                f"{km_inicial:.1f}".replace(".", ",")
                                if km_inicial
                                else ""
                            ),
                            "km_final_mes": f"{km_final:.1f}".replace(".", ","),
                        }
                    )
                    km_anterior_por_veiculo[placa] = km_final

        if not dados_relatorio:
            flash(
                f"Nenhum abastecimento encontrado para os filtros selecionados.",
                "warning",
            )
            return redirect(url_for("selecionar_relatorio_veiculos"))

        output = io.StringIO()
        # A coluna 'capacidade' não existe no dicionário, foi removida do header
        header = [
            "modelo",
            "placa",
            "renavam",
            "ano_fab",
            "ano_mod",
            "tipo_veiculo",
            "capacidade",
            "orgao_localizacao",
            "qtde_abastecimento",
            "combustivel",
            "km_inicial_mes",
            "km_final_mes",
        ]
        writer = csv.DictWriter(output, fieldnames=header, delimiter=";")
        writer.writeheader()

        # O DictWriter espera um dicionário com todas as chaves do cabeçalho
        # Precisamos garantir que todas as chaves existam em cada linha
        rows_completas = []
        for row in dados_relatorio:
            # Adiciona a chave 'capacidade' vazia se ela não existir
            if "capacidade" not in row:
                row["capacidade"] = ""
            rows_completas.append(row)

        writer.writerows(rows_completas)

        response = Response(
            output.getvalue().encode("utf-8-sig"),
            mimetype="text/csv",
            headers={
                "Content-Disposition": f"attachment;filename=relatorio_detalhado_{mes}-{ano}.csv"
            },
        )
        return response

    except Exception as e:
        db.session.rollback()
        flash(f"Ocorreu um erro ao gerar o relatório: {e}", "danger")
        return redirect(url_for("selecionar_relatorio_veiculos"))


@app.route("/usuarios/delete/<int:id>")
@login_required
@admin_required
def delete_usuario(id):
    if User.query.count() <= 1:
        flash("Não é possível excluir o último usuário do sistema.", "danger")
        return redirect(url_for("lista_usuarios"))
    user_to_delete = User.query.get_or_404(id)
    if user_to_delete.username == session.get("username"):
        flash("Você não pode excluir seu próprio usuário.", "danger")
        return redirect(url_for("lista_usuarios"))
    username_deleted = user_to_delete.username
    db.session.delete(user_to_delete)
    db.session.commit()
    registrar_log(f'Excluiu o usuário: "{username_deleted}".')
    flash(f'Usuário "{username_deleted}" excluído.', "success")
    return redirect(url_for("lista_usuarios"))


@app.route("/login", methods=["GET", "POST"])
def login():
    print(f"DEBUG: A aplicação está a usar a base de dados: {app.config['SQLALCHEMY_DATABASE_URI']}")
    if request.method == "POST":
        username = request.form.get("usuario")
        password = request.form.get("senha")
        secretaria_nome_selecionada = request.form.get("secretaria")

        if not secretaria_nome_selecionada:
            flash(
                "Seleção de secretaria inválida. Por favor, reinicie o processo.",
                "danger",
            )
            return redirect(url_for("login"))

        user = User.query.filter_by(username=username).first()

        # Verifica se o usuário e a senha estão corretos primeiro
        if user and bcrypt.check_password_hash(user.password_hash, password):

            if not user.secretaria:
                flash(
                    f"Erro: O usuário '{user.username}' não está associado a nenhuma secretaria. Por favor, contate o administrador.",
                    "danger",
                )
                return redirect(url_for("login"))

            if user.secretaria.nome != secretaria_nome_selecionada:
                flash("Usuário não pertence à secretaria selecionada.", "danger")
                return redirect(url_for("login"))

            # Se tudo estiver correto, continua com o login
            session["logged_in"] = True
            session["username"] = user.username
            session["role"] = user.role
            session["secretaria"] = user.secretaria.nome
            session["secretaria_id"] = user.secretaria.id
            session["user_id"] = user.id  # <-- LINHA ADICIONADA E FORMATADA CORRETAMENTE

            registrar_log(
                f"Fez login no sistema pela secretaria '{user.secretaria.nome}'."
            )
            flash("Login realizado com sucesso!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Usuário ou senha inválidos.", "danger")
            return redirect(url_for("login"))

    # Lógica GET continua a mesma
    secretarias_cadastradas = Secretaria.query.order_by(Secretaria.nome).all()
    secretarias = [s.nome for s in secretarias_cadastradas]
    return render_template("login.html", secretarias=secretarias)


@app.route("/veiculo/<string:placa>/manutencao/add", methods=["POST"])
@login_required
@role_required("Combustivel", "admin")
def add_manutencao(placa):
    veiculo = Veiculo.query.get_or_404(placa)
    try:
        data = datetime.strptime(request.form.get("data_manutencao"), "%Y-%m-%d").date()
        quilometragem = float(
            request.form.get("km_manutencao").replace(".", "").replace(",", ".")
        )
        custo = float(request.form.get("custo_manutencao").replace(",", "."))

        nova_manutencao = Manutencao(
            data=data,
            quilometragem=quilometragem,
            tipo_servico=request.form.get("tipo_servico"),
            custo=custo,
            descricao=request.form.get("descricao"),
            oficina=request.form.get("oficina"),
            veiculo_placa=placa,
        )
        db.session.add(nova_manutencao)
        db.session.commit()
        registrar_log(f"Registrou manutenção de R${custo} para o veículo {placa}.")
        flash("Registro de manutenção salvo com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao salvar manutenção: {e}", "danger")

    return redirect(url_for("detalhes_veiculo", placa=placa))


@app.route("/")
@login_required
def dashboard():
    # Pega a permissão e a secretaria do usuário logado na sessão
    user_role = session.get("role")
    secretaria_id_logada = session.get("secretaria_id")

    # --- LÓGICA DE FILTRO INTERATIVO ---
    # O admin pode passar um filtro de secretaria pela URL (ex: /?secretaria_id=2)
    secretaria_filtro_id = request.args.get("secretaria_id", type=int)

    # Define qual secretaria será usada para filtrar os dados
    target_secretaria_id = secretaria_id_logada

    # Se o usuário é admin E ele selecionou um filtro, a consulta usará o ID do filtro
    if user_role == "admin" and secretaria_filtro_id:
        target_secretaria_id = secretaria_filtro_id

    # --- PREPARAÇÃO DAS QUERIES BASE ---
    query_servidores = Servidor.query
    query_veiculos = Veiculo.query

    # Aplica o filtro de secretaria se o usuário não for admin, ou se o admin aplicou um filtro
    if user_role != "admin" or secretaria_filtro_id:
        query_servidores = query_servidores.filter_by(
            secretaria_id=target_secretaria_id
        )
        query_veiculos = query_veiculos.filter_by(secretaria_id=target_secretaria_id)

    # --- CÁLCULOS PARA OS CARDS E GRÁFICOS ---
    total_servidores = query_servidores.count()
    remuneracao_media_query = query_servidores.with_entities(
        db.func.avg(Servidor.remuneracao)
    ).scalar()
    remuneracao_media = remuneracao_media_query or 0

    # Gráfico de Funções
    servidores_por_funcao = (
        query_servidores.with_entities(Servidor.funcao, db.func.count(Servidor.funcao))
        .group_by(Servidor.funcao)
        .order_by(db.func.count(Servidor.funcao).desc())
        .all()
    )
    funcao_labels = [item[0] or "Não Especificado" for item in servidores_por_funcao]
    funcao_data = [item[1] for item in servidores_por_funcao]

    # Gráfico de Lotações
    servidores_por_lotacao = (
        query_servidores.with_entities(
            Servidor.lotacao, db.func.count(Servidor.lotacao)
        )
        .group_by(Servidor.lotacao)
        .order_by(db.func.count(Servidor.lotacao).desc())
        .all()
    )
    lotacao_labels = [item[0] or "Não Especificado" for item in servidores_por_lotacao]
    lotacao_data = [item[1] for item in servidores_por_lotacao]

    # --- LÓGICA PARA COMBUSTÍVEL ---
    ano_atual = datetime.now().year

    placas_permitidas = [v.placa for v in query_veiculos.all()]
    total_litros_ano = 0
    if placas_permitidas:
        total_litros_ano = (
            db.session.query(db.func.sum(Abastecimento.litros))
            .filter(
                Abastecimento.veiculo_placa.in_(placas_permitidas),
                db.extract("year", Abastecimento.data) == ano_atual,
            )
            .scalar()
            or 0
        )

    dados_grafico_combustivel = []
    if placas_permitidas:
        dados_grafico_combustivel = (
            db.session.query(
                db.func.extract("month", Abastecimento.data),
                db.func.sum(Abastecimento.litros),
            )
            .filter(
                Abastecimento.veiculo_placa.in_(placas_permitidas),
                db.extract("year", Abastecimento.data) == ano_atual,
            )
            .group_by(db.func.extract("month", Abastecimento.data))
            .order_by(db.func.extract("month", Abastecimento.data))
            .all()
        )

    meses_labels = [
        "Jan",
        "Fev",
        "Mar",
        "Abr",
        "Mai",
        "Jun",
        "Jul",
        "Ago",
        "Set",
        "Out",
        "Nov",
        "Dez",
    ]
    litros_data = [0] * 12
    for mes, total in dados_grafico_combustivel:
        if mes is not None:
            litros_data[int(mes) - 1] = total

    # --- LÓGICA PARA ALERTAS ---
    hoje = datetime.now().date()
    data_limite = hoje + timedelta(days=60)
    contratos_a_vencer = (
        query_servidores.filter(
            Servidor.data_saida.isnot(None),
            Servidor.data_saida >= hoje,
            Servidor.data_saida <= data_limite,
        )
        .order_by(Servidor.data_saida.asc())
        .all()
    )
    servidores_incompletos = (
        query_servidores.filter(
            or_(
                Servidor.cpf.is_(None),
                Servidor.cpf == "",
                Servidor.rg.is_(None),
                Servidor.rg == "",
            )
        )
        .order_by(Servidor.nome)
        .all()
    )

    # --- DADOS PARA O FILTRO DO ADMIN ---
    secretarias = None
    secretaria_selecionada = None
    if user_role == "admin":
        secretarias = Secretaria.query.order_by(Secretaria.nome).all()
        if target_secretaria_id:
            secretaria_selecionada = Secretaria.query.get(target_secretaria_id)

    # --- ENVIO DE TODAS AS VARIÁVEIS PARA O TEMPLATE ---
    return render_template(
        "dashboard.html",
        total_servidores=total_servidores,
        remuneracao_media=remuneracao_media,
        funcao_labels=funcao_labels,
        funcao_data=funcao_data,
        lotacao_labels=lotacao_labels,
        lotacao_data=lotacao_data,
        total_litros_ano=total_litros_ano,
        combustivel_labels=meses_labels,
        combustivel_data=litros_data,
        secretarias=secretarias,
        secretaria_selecionada=secretaria_selecionada,
        contratos_a_vencer=contratos_a_vencer,
        servidores_incompletos=servidores_incompletos,
        ano_atual=ano_atual,
    )


@app.route("/")
def splash_screen():
    """Renderiza a tela de animação inicial."""
    return render_template("splash.html")


@app.route("/logout")
@login_required
def logout():
    registrar_log("Fez logout do sistema.")
    session.clear()
    flash("Você saiu do sistema.", "info")
    return redirect(url_for("login"))


@app.route("/documentos/upload/<path:servidor_id>", methods=["POST"])
@login_required
@role_required("RH", "admin")
def upload_documento(servidor_id):
    servidor = Servidor.query.get_or_404(servidor_id)
    if "documento" not in request.files:
        flash("Nenhum arquivo enviado.", "danger")
        return redirect(url_for("editar_servidor", id=servidor_id))

    file = request.files["documento"]
    description = request.form.get("descricao")

    if file.filename == "" or not description:
        flash("A descrição e o arquivo são obrigatórios.", "warning")
        return redirect(url_for("editar_servidor", id=servidor_id))

    if file:
        try:
            # --- UPLOAD PARA SUPABASE ---
            # Envia para a pasta 'documentos_servidores' no Supabase
            url_doc = upload_arquivo_para_nuvem(file, pasta="documentos_servidores")
            
            if url_doc:
                # Salva o LINK DA NUVEM no banco, em vez do nome do arquivo
                novo_documento = Documento(
                    filename=url_doc, 
                    description=description, 
                    servidor_id=servidor_id
                )
                db.session.add(novo_documento)
                db.session.commit()

                registrar_log(f'Anexou o documento "{description}" para o servidor "{servidor.nome}".')
                flash("Documento anexado na nuvem com sucesso!", "success")
            else:
                flash("Erro ao enviar documento para a nuvem.", "danger")

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao anexar documento: {e}", "danger")

    return redirect(url_for("editar_servidor", id=servidor_id))


# ==========================================================
# ROTAS DO MÓDULO GAM
# ==========================================================
@app.route("/gam")
@login_required
@role_required("RH", "admin")
def listar_gams():
    gams = GAM.query.order_by(GAM.data_emissao.desc()).all()
    return render_template("listar_gams.html", gams=gams)


@app.route("/gam/novo", methods=["GET", "POST"])
@login_required
@role_required("RH", "admin")
def criar_gam():
    if request.method == "POST":
        servidor_contrato = request.form.get("servidor_num_contrato")
        servidor = Servidor.query.get(servidor_contrato)
        if not servidor:
            flash("Erro: Servidor selecionado é inválido.", "danger")
            return redirect(url_for("criar_gam"))
        data_laudo_str = request.form.get("data_laudo")
        nova_guia = GAM(
            servidor_num_contrato=servidor_contrato,
            texto_inicial_observacoes=request.form.get("texto_inicial_observacoes"),
            data_laudo=(
                datetime.strptime(data_laudo_str, "%Y-%m-%d").date()
                if data_laudo_str
                else None
            ),
            medico_laudo=request.form.get("medico_laudo"),
            dias_afastamento_laudo=request.form.get("dias_afastamento_laudo", type=int),
            justificativa_laudo=request.form.get("justificativa_laudo"),
            cid10=request.form.get("cid10"),
            status="Emitida",
        )
        db.session.add(nova_guia)
        db.session.commit()
        flash("Guia de Atendimento Médico (GAM) gerada com sucesso!", "success")
        return redirect(url_for("listar_gams"))
    servidores_efetivos = (
        Servidor.query.filter(Servidor.tipo_vinculo.ilike("%efetivo%"))
        .order_by(Servidor.nome)
        .all()
    )
    return render_template("gam_form.html", servidores=servidores_efetivos, gam=None)


@app.route("/gam/editar/<int:gam_id>", methods=["GET", "POST"])
@login_required
@role_required("RH", "admin")
def editar_gam(gam_id):
    guia = GAM.query.get_or_404(gam_id)
    if request.method == "POST":
        try:
            guia.texto_inicial_observacoes = request.form.get(
                "texto_inicial_observacoes"
            )
            data_laudo_str = request.form.get("data_laudo")
            guia.data_laudo = (
                datetime.strptime(data_laudo_str, "%Y-%m-%d").date()
                if data_laudo_str
                else None
            )
            guia.medico_laudo = request.form.get("medico_laudo")
            guia.dias_afastamento_laudo = request.form.get(
                "dias_afastamento_laudo", type=int
            )
            guia.justificativa_laudo = request.form.get("justificativa_laudo")
            guia.cid10 = request.form.get("cid10")
            db.session.commit()
            flash("Guia atualizada com sucesso!", "success")
            return redirect(url_for("listar_gams"))
        except Exception as e:
            db.session.rollback()
            flash(f"Ocorreu um erro ao atualizar a guia: {e}", "danger")
    servidores_efetivos = (
        Servidor.query.filter(Servidor.tipo_vinculo.ilike("%efetivo%"))
        .order_by(Servidor.nome)
        .all()
    )
    return render_template("gam_form.html", gam=guia, servidores=servidores_efetivos)


@app.route("/gam/excluir/<int:gam_id>")
@login_required
@admin_required
@role_required("RH", "admin")
def excluir_gam(gam_id):
    guia = GAM.query.get_or_404(gam_id)
    try:
        db.session.delete(guia)
        db.session.commit()
        flash("Guia excluída com sucesso.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir a guia: {e}", "danger")
    return redirect(url_for("listar_gams"))


@app.route("/gam/imprimir/<int:gam_id>")
@login_required
@role_required("RH", "admin")
def imprimir_gam(gam_id):
    guia = GAM.query.get_or_404(gam_id)
    pdf_buffer = gerar_pdf_gam(guia)
    response = make_response(pdf_buffer.getvalue())
    response.headers["Content-Type"] = "application/pdf"
    nome_arquivo = f'GAM_{guia.servidor.nome.replace(" ", "_")}_{guia.id}.pdf'
    response.headers["Content-Disposition"] = f"inline; filename={nome_arquivo}"
    return response


def gerar_pdf_gam(guia):
    buffer = io.BytesIO()
    doc = BaseDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=4.5 * cm,
        bottomMargin=1.5 * cm,
    )

    def cabecalho_completo(canvas, doc):
        canvas.saveState()
        try:
            timbre_path = os.path.join(basedir, "static", "timbre.jpg")
            if os.path.exists(timbre_path):
                logo = Image(
                    timbre_path, width=17 * cm, height=2.2 * cm, hAlign="CENTER"
                )
                logo.drawOn(canvas, 2 * cm, A4[1] - 3 * cm)
        except Exception as e:
            print(f"Erro ao carregar o timbre: {e}")
        posicao_y_texto = A4[1] - 3.7 * cm
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(
            10.5 * cm, posicao_y_texto, "GUIA PARA ATENDIMENTO MÉDICO - GAM"
        )
        canvas.setFont("Helvetica-Bold", 11)
        canvas.drawCentredString(
            10.5 * cm,
            posicao_y_texto - 0.5 * cm,
            "PREFEITURA MUNICIPAL DE VALENÇA DO PIAUÍ",
        )
        canvas.restoreState()

    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
    template = PageTemplate(
        id="main_template", frames=[frame], onPage=cabecalho_completo
    )
    doc.addPageTemplates([template])
    story = []
    styles = getSampleStyleSheet()
    style_corpo = ParagraphStyle(
        name="Corpo",
        fontName="Helvetica",
        fontSize=10.5,
        alignment=TA_JUSTIFY,
        leading=13,
    )
    style_negrito = ParagraphStyle(
        name="Negrito", parent=style_corpo, fontName="Helvetica-Bold"
    )
    style_assinatura = ParagraphStyle(
        name="Assinatura", fontName="Helvetica", fontSize=9, alignment=TA_CENTER
    )
    dados_servidor_texto = f"<b>1-NOME DO SERVIDOR:</b> {guia.servidor.nome.upper()} &nbsp;&nbsp; <b>MATRÍCULA N.</b> {guia.servidor.num_contrato} &nbsp;&nbsp; <b>LOTAÇÃO:</b> {guia.servidor.lotacao}"
    if guia.data_laudo:
        locale.setlocale(locale.LC_TIME, "pt_BR.UTF-8")
        data_laudo_formatada = guia.data_laudo.strftime("%d de %B de %Y")
    else:
        data_laudo_formatada = "[data não informada]"
    texto_observacoes = f"{guia.texto_inicial_observacoes or ''} Na declaração médica, datada do dia {data_laudo_formatada}, o(a) médico(a) {guia.medico_laudo or '[médico]'} recomenda {guia.dias_afastamento_laudo or '[dias]'} dias de afastamento das suas atividades laborais, pois declara que {guia.justificativa_laudo or '[justificativa]'}"
    cid_texto = f"<b>CID 10: {guia.cid10 or 'Não informado'}</b>"
    encaminhamento_texto = f"Desse modo, encaminho o(a) servidor(a) {guia.servidor.nome} para perícia médica do município."
    observacoes_data = [
        [
            Paragraph(
                "OBSERVAÇÕES DA CHEFIA (ESPECIFICAR A DESCRIÇÃO DO CARGO/FUNÇÃO)",
                style_negrito,
            )
        ],
        [Paragraph(texto_observacoes, style_corpo)],
        [Paragraph(cid_texto, style_corpo)],
        [Paragraph(encaminhamento_texto, style_corpo)],
    ]
    tabela_observacoes = Table(observacoes_data, colWidths=[18 * cm])
    tabela_observacoes.setStyle(
        TableStyle(
            [("BOTTOMPADDING", (0, 0), (0, 0), 8), ("TOPPADDING", (0, 1), (0, -1), 8)]
        )
    )
    data_hora_texto = f"DATA: {guia.data_emissao.strftime('%d/%m/%Y')}, HORA: {guia.data_emissao.strftime('%H:%M')}"
    assinatura_chefia_data = [
        [Paragraph(data_hora_texto, style_corpo)],
        [Spacer(1, 1 * cm)],
        [HRFlowable(width="70%", thickness=0.5, color=colors.black, hAlign="CENTER")],
        [Paragraph("ASSINATURA E CARIMBO DA CHEFIA", style_assinatura)],
    ]
    tabela_assinatura_chefia = Table(assinatura_chefia_data, colWidths=[18 * cm])
    medico_data_hora = "<b>2-DATA DO ATENDIMENTO:</b> ______/______/__________ &nbsp;&nbsp;&nbsp; <b>HORA DO ATENDIMENTO:</b>_____:_____ h."
    assinatura_medico_data = [
        [Paragraph(medico_data_hora, style_corpo)],
        [Spacer(1, 0.5 * cm)],
        [Paragraph("<b>RECOMENDAÇÕES DO MÉDICO</b>", style_corpo)],
        [Spacer(1, 2.5 * cm)],
        [Paragraph("DATA: ______/______/__________", style_corpo)],
        [Spacer(1, 1 * cm)],
        [HRFlowable(width="70%", thickness=0.5, color=colors.black, hAlign="CENTER")],
        [Paragraph("ASSINATURA E CARIMBO DO MÉDICO", style_assinatura)],
    ]
    tabela_medico = Table(assinatura_medico_data, colWidths=[18 * cm])
    tabela_principal_data = [
        [Paragraph(dados_servidor_texto, style_corpo)],
        [tabela_observacoes],
        [tabela_assinatura_chefia],
        [HRFlowable(width="100%", thickness=1, color=colors.black, dash=(2, 2))],
        [tabela_medico],
    ]
    tabela_principal = Table(
        tabela_principal_data,
        rowHeights=[1.2 * cm, 7.5 * cm, 3 * cm, 0.3 * cm, 6.5 * cm],
    )
    tabela_principal.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(tabela_principal)
    doc.build(story)
    buffer.seek(0)
    return buffer


@app.route("/api/servidor-details/<string:num_contrato>")
@login_required
@role_required("RH", "admin")
def get_servidor_details(num_contrato):
    servidor = Servidor.query.filter_by(num_contrato=num_contrato).first()
    if servidor:
        return jsonify(
            {
                "nome": servidor.nome,
                "matricula": servidor.num_contrato,
                "lotacao": servidor.lotacao,
                "funcao": servidor.funcao,
                "data_admissao": (
                    servidor.data_inicio.strftime("%d/%m/%Y")
                    if servidor.data_inicio
                    else ""
                ),
            }
        )
    return jsonify({"error": "Servidor não encontrado"}), 404


@app.route("/servidores")
@login_required
@role_required("RH", "admin")
def lista_servidores():
    user_role = session.get("role")
    secretaria_id_logada = session.get("secretaria_id")

    # 1. Filtro base de permissão (Admin vê tudo, RH vê só sua secretaria)
    if user_role == 'admin':
        query = Servidor.query
    else:
        query = Servidor.query.filter_by(secretaria_id=secretaria_id_logada)

    # 2. Captura os filtros da URL (busca, função, lotação)
    termo_busca = request.args.get("termo")
    funcao_filtro = request.args.get("funcao")
    lotacao_filtro = request.args.get("lotacao")

    # 3. Aplica os filtros na Query
    if termo_busca:
        # Cria padrão de busca parcial (ilike)
        search_pattern = f"%{termo_busca}%"
        query = query.filter(
            or_(
                Servidor.nome.ilike(search_pattern),
                Servidor.cpf.ilike(search_pattern),
                Servidor.num_contrato.ilike(search_pattern),
            )
        )

    if funcao_filtro:
        query = query.filter(Servidor.funcao == funcao_filtro)

    if lotacao_filtro:
        query = query.filter(Servidor.lotacao == lotacao_filtro)

    # 4. Executa a busca final
    servidores = query.order_by(Servidor.nome).all()

    # 5. Prepara listas para os dropdowns de filtro
    funcoes_disponiveis = [r[0] for r in db.session.query(Servidor.funcao).distinct().order_by(Servidor.funcao).all() if r[0]]
    lotacoes_disponiveis = [r[0] for r in db.session.query(Servidor.lotacao).distinct().order_by(Servidor.lotacao).all() if r[0]]

    # 6. IMPORTANTE: Busca escolas para o modal de cadastro novo
    escolas = Escola.query.filter_by(status='Ativa').order_by(Escola.nome).all()

    # 7. Verifica status (férias/licença)
    hoje = datetime.now().date()
    status_servidores = {}
    requerimentos_ativos = Requerimento.query.filter(
        Requerimento.status == "Aprovado", Requerimento.data_inicio_requerimento <= hoje
    ).all()

    for req in requerimentos_ativos:
        if not req.data_retorno_trabalho or req.data_retorno_trabalho > hoje:
            status_servidores[req.servidor_cpf] = req.natureza

    # 8. Renderiza o template enviando TUDO
    return render_template(
        "index.html",
        servidores=servidores,
        funcoes_disponiveis=funcoes_disponiveis,
        lotacoes_disponiveis=lotacoes_disponiveis,
        status_servidores=status_servidores,
        escolas=escolas  # <--- Essencial para o cadastro funcionar
    )


@app.route("/delete/<path:id>")
@login_required
@admin_required
@role_required("RH", "admin")
def delete_server(id):
    servidor = Servidor.query.get_or_404(id)
    dependencias = []

    if hasattr(servidor, "contratos") and servidor.contratos:
        dependencias.append("contratos")
    if hasattr(servidor, "requerimentos") and servidor.requerimentos:
        dependencias.append("requerimentos")
    if hasattr(servidor, "pontos") and servidor.pontos:
        dependencias.append("registros de ponto")

    if dependencias:
        dependencias_str = ", ".join(dependencias)
        flash(
            f'Não é possível excluir o servidor "{servidor.nome}", pois ele possui vínculos com: {dependencias_str}.',
            "danger",
        )
        return redirect(url_for("lista_servidores"))

    nome_servidor = servidor.nome
    try:
        if servidor.foto_filename and os.path.exists(
            os.path.join(app.config["UPLOAD_FOLDER"], servidor.foto_filename)
        ):
            os.remove(os.path.join(app.config["UPLOAD_FOLDER"], servidor.foto_filename))
        for doc in servidor.documentos:
            doc_path = os.path.join(
                app.config["UPLOAD_FOLDER"], "documentos", doc.filename
            )
            if os.path.exists(doc_path):
                os.remove(doc_path)

        db.session.delete(servidor)
        db.session.commit()
        registrar_log(f'Excluiu o servidor: "{nome_servidor}".')
        flash(f'Servidor "{nome_servidor}" excluído com sucesso!', "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Ocorreu um erro ao tentar excluir o servidor: {e}", "danger")

    return redirect(url_for("lista_servidores"))


@app.route("/importar_servidores", methods=["POST"])
@login_required
@admin_required
@role_required("RH", "admin")
def importar_servidores():
    if "csv_file" not in request.files:
        flash("Nenhum arquivo enviado.", "danger")
        return redirect(url_for("lista_servidores"))

    file = request.files["csv_file"]
    if file.filename == "":
        flash("Nenhum arquivo selecionado.", "danger")
        return redirect(url_for("lista_servidores"))

    if file and file.filename.endswith(".csv"):
        try:
            # Lógica para ler e processar o arquivo CSV
            # (Esta parte do seu código original)
            db.session.commit()
            flash("Importação concluída com sucesso!", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Ocorreu um erro ao processar o arquivo: {e}", "danger")

        return redirect(url_for("lista_servidores"))
    else:
        flash(
            "Formato de arquivo inválido. Por favor, envie um arquivo .csv.", "warning"
        )
        return redirect(url_for("lista_servidores"))


@app.route("/baixar_modelo_csv")
@login_required
def baixar_modelo_csv():
    header = [
        "Nº CONTRATO",
        "NOME",
        "FUNÇÃO",
        "LOTAÇÃO",
        "CARGA HORÁRIA",
        "REMUNERAÇÃO",
        "VIGÊNCIA",
    ]
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(header)
    csv_content = output.getvalue()

    response = Response(
        csv_content.encode("utf-8-sig"),
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment;filename=modelo_importacao_servidores.csv"
        },
    )
    return response


@app.route("/veiculo/<string:placa>/detalhes")
@login_required
@check_license
@role_required("Combustivel", "admin")
def detalhes_veiculo(placa):
    veiculo = Veiculo.query.get_or_404(placa)
    abastecimentos = (
        Abastecimento.query.filter_by(veiculo_placa=placa)
        .order_by(Abastecimento.quilometragem.asc())
        .all()
    )
    manutencoes = (
        Manutencao.query.filter_by(veiculo_placa=placa)
        .order_by(Manutencao.data.desc())
        .all()
    )
    indicadores = {
        "gasto_combustivel": sum(a.valor_total for a in abastecimentos),
        "gasto_manutencao": sum(m.custo for m in manutencoes),
        "total_litros": sum(a.litros for a in abastecimentos),
        "total_km_rodado": 0,
        "consumo_medio_geral": 0,
        "custo_medio_km": 0,
    }
    indicadores["gasto_total"] = (
        indicadores["gasto_combustivel"] + indicadores["gasto_manutencao"]
    )
    chart_labels = []
    chart_consumo_data = []
    chart_custo_km_data = []
    abastecimentos_com_analise = []
    if len(abastecimentos) > 1:
        indicadores["total_km_rodado"] = (
            abastecimentos[-1].quilometragem - abastecimentos[0].quilometragem
        )
        if indicadores["total_km_rodado"] > 0:
            litros_para_media = sum(a.litros for a in abastecimentos[:-1])
            if litros_para_media > 0:
                indicadores["consumo_medio_geral"] = (
                    indicadores["total_km_rodado"] / litros_para_media
                )
            if indicadores["gasto_total"] > 0:
                indicadores["custo_medio_km"] = (
                    indicadores["gasto_total"] / indicadores["total_km_rodado"]
                )
        for i in range(1, len(abastecimentos)):
            anterior = abastecimentos[i - 1]
            atual = abastecimentos[i]
            analise = {"abastecimento": atual, "km_rodado": 0, "consumo_kml": 0}
            km_rodado = atual.quilometragem - anterior.quilometragem
            if km_rodado > 0 and anterior.litros > 0:
                consumo_kml = km_rodado / anterior.litros
                custo_km = anterior.valor_total / km_rodado
                analise.update({"km_rodado": km_rodado, "consumo_kml": consumo_kml})
                chart_labels.append(atual.data.strftime("%d/%m"))
                chart_consumo_data.append(round(consumo_kml, 2))
                chart_custo_km_data.append(round(custo_km, 2))
            abastecimentos_com_analise.append(analise)
    abastecimentos_com_analise.reverse()

    return render_template(
        "detalhes_veiculo.html",
        veiculo=veiculo,
        indicadores=indicadores,
        abastecimentos_com_analise=abastecimentos_com_analise,
        manutencoes=manutencoes,
        chart_labels=chart_labels,
        chart_consumo_data=chart_consumo_data,
        chart_custo_km_data=chart_custo_km_data,
    )


@app.route("/add", methods=["POST"])
@login_required
@role_required("RH", "admin")
def add_server():
    try:
        secretaria_id_do_usuario = session.get("secretaria_id")
        if not secretaria_id_do_usuario:
            flash("Erro de sessão. Faça login novamente.", "danger")
            return redirect(url_for("lista_servidores"))

        # 1. Upload da Foto e Biometria
        foto = request.files.get("foto")
        foto_filename = None
        face_encoding_blob = None
        
        if foto and foto.filename != "":
            # Tenta gerar o encoding facial antes de salvar na nuvem
            try:
                # Carrega a imagem para o face_recognition
                image_file = face_recognition.load_image_file(foto)
                encodings = face_recognition.face_encodings(image_file)
                
                if len(encodings) > 0:
                    face_encoding_blob = json.dumps(encodings[0].tolist())
                
                # Volta o ponteiro do arquivo para o início para fazer o upload
                foto.seek(0)
                
                # Envia para o Supabase
                url_foto = upload_arquivo_para_nuvem(foto, pasta="fotos_servidores")
                if url_foto:
                    foto_filename = url_foto
            except Exception as e:
                print(f"Erro ao processar biometria no cadastro: {e}")
                # Não impede o cadastro, mas avisa no log

        # 2. Tratamento de Datas
        data_inicio_str = request.form.get("data_inicio")
        data_saida_str = request.form.get("data_saida")
        data_nascimento_str = request.form.get("data_nascimento")
        
        data_inicio_obj = datetime.strptime(data_inicio_str, "%Y-%m-%d").date() if data_inicio_str else None
        data_saida_obj = datetime.strptime(data_saida_str, "%Y-%m-%d").date() if data_saida_str else None
        data_nascimento_obj = datetime.strptime(data_nascimento_str, "%Y-%m-%d").date() if data_nascimento_str else None
        
        # 3. Limpeza de Dados
        cpf_limpo = limpar_cpf(request.form.get("cpf"))
        remuneracao_str = request.form.get("remuneracao", "0").replace(".", "").replace(",", ".")
        try:
            remuneracao_val = float(remuneracao_str)
        except:
            remuneracao_val = 0.0
        
        # 4. Criação do Objeto Servidor
        novo_servidor = Servidor(
            num_contrato=request.form.get("num_contrato"),
            nome=request.form.get("nome"),
            cpf=cpf_limpo,
            rg=request.form.get("rg"),
            data_nascimento=data_nascimento_obj,
            nome_mae=request.form.get("nome_mae"),
            email=request.form.get("email"),
            pis_pasep=request.form.get("pis_pasep"),
            tipo_vinculo=request.form.get("tipo_vinculo"),
            local_trabalho=request.form.get("local_trabalho"),
            escola_id=request.form.get("escola_id_vinculo", type=int),  # <--- VINCULO SALVO AQUI
            classe_nivel=request.form.get("classe_nivel"),
            num_contra_cheque=request.form.get("num_contra_cheque"),
            nacionalidade=request.form.get("nacionalidade"),
            estado_civil=request.form.get("estado_civil"),
            telefone=request.form.get("telefone"),
            endereco=request.form.get("endereco"),
            funcao=request.form.get("funcao"),
            lotacao=request.form.get("lotacao"),
            carga_horaria=request.form.get("carga_horaria"),
            remuneracao=remuneracao_val,
            dados_bancarios=request.form.get("dados_bancarios"),
            data_inicio=data_inicio_obj,
            data_saida=data_saida_obj,
            observacoes=request.form.get("observacoes"),
            foto_filename=foto_filename,
            face_encoding=face_encoding_blob, # Salva a biometria calculada
            secretaria_id=secretaria_id_do_usuario
        )

        db.session.add(novo_servidor)
        db.session.commit()
        flash("Servidor cadastrado com sucesso!", "success")

    except Exception as e:
        db.session.rollback()
        print(f"Erro no cadastro: {e}")
        flash(f"Erro ao cadastrar: {e}", "danger")

    return redirect(url_for("lista_servidores"))

@app.route("/editar/<path:id>", methods=["GET", "POST"])
@login_required
@role_required("RH", "admin")
def editar_servidor(id):
    servidor = Servidor.query.get_or_404(id)
    secretarias = Secretaria.query.order_by(Secretaria.nome).all()
    
    # 1. IMPORTANTE: Carrega escolas para o dropdown de edição
    escolas = Escola.query.filter_by(status='Ativa').order_by(Escola.nome).all()
    
    if request.method == "POST":
        try:
            # --- Lógica de Foto e Biometria ---
            foto = request.files.get("foto")
            if foto and foto.filename != "":
                # 1. Calcula biometria nova
                try:
                    image_file = face_recognition.load_image_file(foto)
                    encodings = face_recognition.face_encodings(image_file)
                    if len(encodings) > 0:
                        servidor.face_encoding = json.dumps(encodings[0].tolist())
                        flash("Biometria facial atualizada com sucesso!", "info")
                    
                    foto.seek(0) # Reseta ponteiro para upload
                except Exception as e:
                    print(f"Erro ao processar face na edição: {e}")
                
                # 2. Faz upload
                url_foto = upload_arquivo_para_nuvem(foto, pasta="fotos_servidores")
                if url_foto:
                    servidor.foto_filename = url_foto

            # --- Atualização de Dados Cadastrais ---
            servidor.nome = request.form.get("nome")
            servidor.cpf = limpar_cpf(request.form.get("cpf"))
            servidor.rg = request.form.get("rg")
            servidor.email = request.form.get("email")
            servidor.telefone = request.form.get("telefone")
            servidor.endereco = request.form.get("endereco")
            servidor.funcao = request.form.get("funcao")
            servidor.lotacao = request.form.get("lotacao")
            
            # ATUALIZA O VÍNCULO DA ESCOLA
            servidor.escola_id = request.form.get("escola_id_vinculo", type=int)

            servidor.tipo_vinculo = request.form.get("tipo_vinculo")
            servidor.carga_horaria = request.form.get("carga_horaria")
            servidor.dados_bancarios = request.form.get("dados_bancarios")
            servidor.observacoes = request.form.get("observacoes")

            # --- Tratamento de Valores ---
            remuneracao_str = request.form.get("remuneracao", "0").replace(".", "").replace(",", ".")
            try:
                servidor.remuneracao = float(remuneracao_str)
            except:
                pass # Mantém o valor antigo se der erro

            # --- Tratamento de Datas ---
            data_nasc = request.form.get("data_nascimento")
            data_ini = request.form.get("data_inicio")
            data_sai = request.form.get("data_saida")

            if data_nasc:
                servidor.data_nascimento = datetime.strptime(data_nasc, "%Y-%m-%d").date()
            if data_ini:
                servidor.data_inicio = datetime.strptime(data_ini, "%Y-%m-%d").date()
            if data_sai:
                servidor.data_saida = datetime.strptime(data_sai, "%Y-%m-%d").date()
            else:
                servidor.data_saida = None # Permite limpar a data de saída

            db.session.commit()
            flash("Servidor atualizado com sucesso!", "success")
            return redirect(url_for("lista_servidores"))
            
        except Exception as e:
            db.session.rollback()
            print(f"Erro na edição: {e}")
            flash(f"Erro ao atualizar servidor: {e}", "danger")
            return redirect(url_for("editar_servidor", id=id))

    # --- Renderiza enviando 'escolas' ---
    return render_template("editar.html", servidor=servidor, secretarias=secretarias, escolas=escolas)

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """ Rota para servir os arquivos da pasta 'uploads' (como fotos de perfil). """
    # 'send_from_directory' já está importado no topo do seu app.py
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route("/veiculos", methods=["GET", "POST"])
@login_required
@fleet_required
@check_license
@role_required("Combustivel", "admin")
def gerenciar_veiculos():
    # Lógica para CADASTRAR um novo veículo (quando o formulário é enviado)
    if request.method == "POST":
        try:
            nova_placa = request.form.get("placa").upper().strip()
            veiculo_existente = Veiculo.query.get(nova_placa)
            if veiculo_existente:
                flash("Veículo com esta placa já cadastrado.", "danger")
                return redirect(url_for("gerenciar_veiculos"))

            novo_veiculo = Veiculo(
                placa=nova_placa,
                modelo=request.form.get("modelo"),
                tipo=request.form.get("tipo"),
                ano_fabricacao=request.form.get("ano_fabricacao", type=int),
                ano_modelo=request.form.get("ano_modelo", type=int),
                # Salva o ID da secretaria selecionada no formulário
                secretaria_id=request.form.get("secretaria_id", type=int),
                renavam=request.form.get("renavam") or None,
                autorizacao_detran=request.form.get("autorizacao_detran") or None,
                validade_autorizacao=(
                    datetime.strptime(
                        request.form.get("validade_autorizacao"), "%Y-%m-%d"
                    ).date()
                    if request.form.get("validade_autorizacao")
                    else None
                ),
                certificado_tacografo=request.form.get("certificado_tacografo") or None,
                data_emissao_tacografo=(
                    datetime.strptime(
                        request.form.get("data_emissao_tacografo"), "%Y-%m-%d"
                    ).date()
                    if request.form.get("data_emissao_tacografo")
                    else None
                ),
                validade_tacografo=(
                    datetime.strptime(
                        request.form.get("validade_tacografo"), "%Y-%m-%d"
                    ).date()
                    if request.form.get("validade_tacografo")
                    else None
                ),
            )
            db.session.add(novo_veiculo)
            db.session.commit()
            registrar_log(
                f"Cadastrou o veículo: Placa {novo_veiculo.placa}, Modelo {novo_veiculo.modelo}."
            )
            flash("Veículo cadastrado com sucesso!", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao cadastrar veículo: {e}", "danger")

        return redirect(url_for("gerenciar_veiculos"))

    # Lógica para EXIBIR a página de veículos

    # Inicia a consulta base de veículos
    query = Veiculo.query

    # --- LÓGICA DO SUPER ADMIN ---
    # Se o usuário logado NÃO for um admin, aplica o filtro da secretaria
    if session.get("role") != "admin":
        secretaria_id_logada = session.get("secretaria_id")
        query = query.filter_by(secretaria_id=secretaria_id_logada)

    veiculos = query.order_by(Veiculo.modelo).all()

    # Busca a lista de secretarias para popular o formulário de cadastro
    secretarias = Secretaria.query.order_by(Secretaria.nome).all()

    # Lógica para os alertas de documentação
    hoje = datetime.now().date()
    data_limite = hoje + timedelta(days=30)
    veiculos_com_alerta = [
        v
        for v in veiculos
        if (v.validade_autorizacao and v.validade_autorizacao <= data_limite)
        or (v.validade_tacografo and v.validade_tacografo <= data_limite)
    ]

    return render_template(
        "veiculos.html",
        veiculos=veiculos,
        secretarias=secretarias,
        veiculos_com_alerta=veiculos_com_alerta,
        hoje=hoje,
        data_limite=data_limite,
    )


@app.route("/veiculos/excluir/<path:placa>")
@login_required
def excluir_veiculo(placa):
    veiculo = Veiculo.query.get_or_404(placa)
    try:
        db.session.delete(veiculo)
        db.session.commit()
        registrar_log(f"Excluiu o veículo: Placa {veiculo.placa}.")
        flash("Veículo excluído com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Não foi possível excluir o veículo. Erro: {e}", "danger")
    return redirect(url_for("gerenciar_veiculos"))


@app.route("/requerimentos")
@login_required
@role_required("RH", "admin")
def listar_requerimentos():
    requerimentos = Requerimento.query.order_by(Requerimento.data_criacao.desc()).all()
    return render_template("requerimentos.html", requerimentos=requerimentos)


@app.route("/requerimentos/novo", methods=["GET", "POST"])
@login_required
@role_required("RH", "admin")
def novo_requerimento():
    if request.method == "POST":
        try:
            cpf_servidor = limpar_cpf(request.form.get("cpf_busca"))
            servidor = Servidor.query.filter_by(cpf=cpf_servidor).first()
            if not servidor:
                flash("Servidor com o CPF informado não encontrado.", "danger")
                return redirect(url_for("novo_requerimento"))

            novo_req = Requerimento(
                autoridade_dirigida=request.form.get("autoridade_dirigida"),
                servidor_cpf=cpf_servidor,
                natureza=request.form.get("natureza"),
                data_inicio_requerimento=datetime.strptime(
                    request.form.get("data_inicio_requerimento"), "%Y-%m-%d"
                ).date(),
                status="Em Análise",
                # Adicione outros campos do formulário aqui se necessário
            )
            db.session.add(novo_req)
            db.session.commit()
            flash("Requerimento criado com sucesso!", "success")
            return redirect(url_for("listar_requerimentos"))
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao criar requerimento: {e}", "danger")
    return render_template("requerimento_form.html")

@app.route("/requerimentos/excluir/<int:req_id>")
@login_required
@role_required("RH", "admin")
def excluir_requerimento(req_id):
    requerimento = Requerimento.query.get_or_404(req_id)
    try:
        # Registra quem excluiu para segurança
        registrar_log(f'Excluiu o requerimento #{requerimento.id} do servidor {requerimento.servidor.nome}.')
        
        db.session.delete(requerimento)
        db.session.commit()
        flash("Requerimento excluído com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir requerimento: {e}", "danger")

    return redirect(url_for("listar_requerimentos"))


@app.route("/requerimento/pdf/<int:req_id>")
@login_required
@role_required('RH', 'admin')
def gerar_requerimento_pdf(req_id):
    requerimento = Requerimento.query.get_or_404(req_id)
    servidor = requerimento.servidor

    def fmt_data(data):
        return data.strftime('%d/%m/%Y') if data else ""

    data_hoje = datetime.now().strftime('%d de %B de %Y')
    
    natureza_texto = requerimento.natureza
    if requerimento.natureza == 'Outro' and requerimento.natureza_outro:
        natureza_texto = f"{requerimento.natureza} ({requerimento.natureza_outro})"

    buffer = io.BytesIO()
    
    # --- AJUSTE 1: Margens Verticais Otimizadas ---
    # Top 4.5cm para respeitar o timbre, mas Bottom reduzido para 1.5cm para caber mais texto
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4, 
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=4.5*cm,    
        bottomMargin=1.5*cm  
    )
    
    largura_util = 18 * cm
    
    styles = getSampleStyleSheet()
    # Fonte ligeiramente menor (8.5) para economizar espaço
    style_normal = ParagraphStyle('Normal_Custom', parent=styles['Normal'], fontSize=8.5, leading=10)
    style_center = ParagraphStyle('Center_Custom', parent=styles['Normal'], fontSize=8.5, alignment=TA_CENTER, leading=10)
    style_center_bold = ParagraphStyle('Center_Bold', parent=styles['Normal'], fontSize=8.5, alignment=TA_CENTER, fontName='Helvetica-Bold', leading=10)
    style_title = ParagraphStyle('Title_Custom', parent=styles['Normal'], fontSize=11, alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=5)
    
    def label(texto):
        return Paragraph(f"<b>{texto}</b>", style_normal)
    
    def content(texto):
        return Paragraph(str(texto or ""), style_normal)

    story = []

    # TÍTULO
    story.append(Paragraph("REQUERIMENTO PADRÃO", style_title))
    # Espaço reduzido
    story.append(Spacer(1, 0.2*cm))

    # DESTINATÁRIO
    tbl_destinatario_data = [
        [label("AUTORIDADE A QUEM É DIRIGIDA:"), content(requerimento.autoridade_dirigida or "Sr(a). Secretário(a)")]
    ]
    tbl_destinatario = Table(tbl_destinatario_data, colWidths=[7*cm, 11*cm])
    tbl_destinatario.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2), # Padding reduzido
    ]))
    story.append(tbl_destinatario)
    story.append(Spacer(1, 0.3*cm))

    # BLOCO 1: IDENTIFICAÇÃO
    story.append(Paragraph("<b>1. IDENTIFICAÇÃO DO SERVIDOR</b>", style_normal))
    story.append(Spacer(1, 0.1*cm))
    
    cols_width = [4*cm, 5*cm, 4*cm, 5*cm]

    dados_servidor = [
        [label("NOME COMPLETO:"), content(servidor.nome), label("Nº MATRÍCULA:"), content(servidor.num_contrato)],
        [label("CARGO/FUNÇÃO:"), content(servidor.funcao), label("CLASSE/NÍVEL:"), content(servidor.classe_nivel)],
        [label("DATA NASCIMENTO:"), content(fmt_data(servidor.data_nascimento)), label("DATA ADMISSÃO:"), content(fmt_data(servidor.data_inicio))],
        [label("LOTAÇÃO:"), content(servidor.lotacao), label("TELEFONE:"), content(servidor.telefone)],
        [label("LOCAL DE TRABALHO:"), content(servidor.local_trabalho), label(""), content("")],
        [label("ENDEREÇO:"), content(servidor.endereco), label(""), content("")]
    ]

    t_servidor = Table(dados_servidor, colWidths=cols_width)
    t_servidor.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('SPAN', (1, 5), (3, 5)),
        ('SPAN', (1, 4), (3, 4)),
        ('BACKGROUND', (0,0), (0,-1), colors.whitesmoke),
        ('BACKGROUND', (2,0), (2,3), colors.whitesmoke),
        # Padding reduzido para 2.5 para economizar altura
        ('TOPPADDING', (0,0), (-1,-1), 2.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2.5),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(t_servidor)
    story.append(Spacer(1, 0.3*cm))

    # BLOCO 2: DADOS DO REQUERIMENTO
    story.append(Paragraph("<b>2. DADOS DO REQUERIMENTO</b>", style_normal))
    story.append(Spacer(1, 0.1*cm))

    dados_req = [
        [label("NATUREZA:"), content(natureza_texto), label("DATA INÍCIO:"), content(fmt_data(requerimento.data_inicio_requerimento))],
        [label("PERÍODO AQUISITIVO:"), content(requerimento.periodo_aquisitivo), label("DURAÇÃO (DIAS):"), content(requerimento.duracao)],
        [label("INFORMAÇÕES COMPLEMENTARES:"), content(requerimento.informacoes_complementares), label(""), content("")]
    ]

    t_req = Table(dados_req, colWidths=cols_width)
    t_req.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('SPAN', (1, 2), (3, 2)),
        ('BACKGROUND', (0,0), (0,-1), colors.whitesmoke),
        ('BACKGROUND', (2,0), (2,1), colors.whitesmoke),
        ('VALIGN', (0, 2), (-1, 2), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 2.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2.5),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(t_req)
    story.append(Spacer(1, 0.3*cm))

    # BLOCO 3: PARECER JURÍDICO
    # Usamos KeepTogether para tentar manter o título e o conteúdo juntos
    bloco_parecer = []
    bloco_parecer.append(Paragraph("<b>3. PARECER JURÍDICO / ADMINISTRATIVO</b>", style_normal))
    bloco_parecer.append(Spacer(1, 0.1*cm))
    
    tbl_parecer_data = [[content(requerimento.parecer_juridico or " ")]]
    
    t_parecer = Table(tbl_parecer_data, colWidths=[largura_util])
    t_parecer.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
        ('MINROWHEIGHT', (0,0), (-1,-1), 1.5*cm)
    ]))
    bloco_parecer.append(t_parecer)
    bloco_parecer.append(Spacer(1, 0.3*cm))
    story.append(KeepTogether(bloco_parecer))

    # ASSINATURA REQUERENTE
    # Criamos um bloco protegido para que a assinatura não fique sozinha na próx página
    bloco_assinatura_req = []
    bloco_assinatura_req.append(Paragraph(f"Valença do Piauí, {data_hoje}", style_center))
    bloco_assinatura_req.append(Spacer(1, 0.8*cm)) # Espaço para assinar
    bloco_assinatura_req.append(Paragraph("________________________________________________________", style_center))
    bloco_assinatura_req.append(Paragraph(f"<b>{servidor.nome.upper()}</b>", style_center))
    bloco_assinatura_req.append(Paragraph("ASSINATURA DO REQUERENTE", style_center))
    bloco_assinatura_req.append(Spacer(1, 0.3*cm))
    bloco_assinatura_req.append(HRFlowable(width="100%", thickness=0.5, color=colors.black, dash=(3, 2)))
    bloco_assinatura_req.append(Spacer(1, 0.3*cm))
    story.append(KeepTogether(bloco_assinatura_req))

    # BLOCO 4: CHEFIA E FINAL
    # Agrupamos Chefia e Assinaturas Finais para tentar manter no fim da página 1
    bloco_final = []
    bloco_final.append(Paragraph("<b>4. DESPACHO DA CHEFIA IMEDIATA</b>", style_normal))
    bloco_final.append(Spacer(1, 0.1*cm))
    
    check_box = " (   ) LIBERADO      (   ) NÃO LIBERADO"
    
    tbl_chefia_data = [
        [Paragraph(check_box, style_normal)],
        [Spacer(1, 0.4*cm)], 
        [Paragraph("_____________________________________________", style_center)],
        [Paragraph("ASSINATURA E CARIMBO DO CHEFE IMEDIATO", style_center)]
    ]
    
    t_chefia = Table(tbl_chefia_data, colWidths=[largura_util])
    t_chefia.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.black),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    bloco_final.append(t_chefia)
    bloco_final.append(Spacer(1, 0.3*cm))

    bloco_final.append(Paragraph("Encaminhe-se ao setor competente para as providências necessárias.", style_normal))
    bloco_final.append(Spacer(1, 0.8*cm))

    col_width_assinatura = largura_util / 2
    tbl_assinaturas_finais = [
        [
            Paragraph("_________________________", style_center),
            Paragraph("_________________________", style_center)
        ],
        [
            Paragraph("SECRETÁRIO(A) MUNICIPAL", style_center_bold),
            Paragraph("SETOR DE RH / PESSOAL", style_center_bold)
        ]
    ]
    
    t_final = Table(tbl_assinaturas_finais, colWidths=[col_width_assinatura, col_width_assinatura])
    t_final.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    bloco_final.append(t_final)
    
    # Adiciona o bloco final protegido
    story.append(KeepTogether(bloco_final))

    doc.build(story, onFirstPage=cabecalho_e_rodape, onLaterPages=cabecalho_e_rodape)
    buffer.seek(0)

    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    nome_arquivo = f'requerimento_{servidor.nome.replace(" ", "_")}_{req_id}.pdf'
    response.headers['Content-Disposition'] = f'inline; filename={nome_arquivo}'
    
    return response

def converter_dias_para_texto(total_dias):
    """
    Função auxiliar para converter um número de dias em Anos, Meses e Dias.
    Usa a lógica de (Ano=365 dias, Mês=30 dias).
    """
    if total_dias < 0:
        total_dias = 0
    anos = total_dias // 365
    meses = (total_dias % 365) // 30
    dias = (total_dias % 365) % 30
    return f"{anos} anos, {meses} meses e {dias} dias"

def calcular_efetivo_exercicio(cpf_servidor):
    """
    Função auxiliar para calcular o tempo de efetivo exercício.
    VERSÃO 4.0: Usa a busca exata (filter_by) que sabemos que funciona.
    """
    
    # cpf_servidor já chega aqui limpo (ex: "06184477331")
    
    # --- [ LINHA ATUALIZADA (v4.0) ] ---
    # Voltando para a busca exata, que é usada em outras partes do seu app.
    # Isso assume que os CPFs no banco de dados estão LIMPOS (sem pontos/traços).
    servidor = Servidor.query.filter_by(cpf=cpf_servidor).first()
    # --- [ FIM DA ATUALIZAÇÃO ] ---

    if not servidor:
        # Se esta busca falha, o CPF limpo não existe na coluna Servidor.cpf
        raise Exception("Servidor não encontrado com este CPF.")
        
    if not servidor.data_inicio:
        raise Exception("Servidor não possui data de admissão (data_inicio) cadastrada.")

    data_admissao = servidor.data_inicio
    data_hoje = datetime.now().date()
    
    # 1. TEMPO BRUTO
    dias_totais_bruto = (data_hoje - data_admissao).days
    
    # --- [ LÓGICA CONDICIONAL DE DESCONTO ] ---
    NATUREZAS_DESCONTO_SEMPRE = [
        "Licença para Tratar de Interesse Particular",
        "Licença por Motivo de Doença em Pessoa da Família"
    ]
    NATUREZAS_DESCONTO_PROFESSOR = [
        "Licença Prêmio"
    ]

    funcao_servidor = (servidor.funcao or "").lower()
    e_professor = "professor" in funcao_servidor or "professora" in funcao_servidor

    lista_descontos_final = list(NATUREZAS_DESCONTO_SEMPRE)
    
    if e_professor:
        lista_descontos_final.extend(NATUREZAS_DESCONTO_PROFESSOR)
        print(f"DEBUG (v4.0): Servidor {servidor.nome} (Função: {servidor.funcao}) é PROFESSOR. Descontando Licença Prêmio.")
    else:
        print(f"DEBUG (v4.0): Servidor {servidor.nome} (Função: {servidor.funcao}) NÃO é professor. Licença Prêmio NÃO será descontada.")

    # 2. TEMPO DE DESCONTO
    # Esta busca também usa o CPF limpo (cpf_servidor)
    afastamentos_requerimentos = Requerimento.query.filter(
        Requerimento.servidor_cpf == cpf_servidor, 
        Requerimento.status == 'Aprovado',
        Requerimento.natureza.in_(lista_descontos_final) 
    ).all()

    dias_afastamento_total = 0
    afastamentos_info = []
    
    for req in afastamentos_requerimentos: 
        data_fim_req = req.data_retorno_trabalho or req.data_conclusao
        if not data_fim_req:
             data_fim_req = data_hoje
             
        if req.data_inicio_requerimento:
             dias_descontados = (data_fim_req - req.data_inicio_requerimento).days
             if dias_descontados > 0:
                 dias_afastamento_total += dias_descontados
                 afastamentos_info.append({
                     "id": req.id,
                     "natureza": req.natureza,
                     "data_inicio": req.data_inicio_requerimento,
                     "data_fim": data_fim_req,
                     "dias": dias_descontados
                 })

    # 3. TEMPO LÍQUIDO
    dias_efetivos = dias_totais_bruto - dias_afastamento_total
    
    return servidor, dias_totais_bruto, dias_afastamento_total, dias_efetivos, afastamentos_info

@app.route("/api/calculo_efetivo_exercicio/<string:cpf>")
@login_required
@role_required("RH", "admin")
def api_calculo_efetivo_exercicio(cpf):
    """
    (v4.0) Adiciona um log de erro detalhado.
    """
    try:
        # A função limpar_cpf está no seu app.py e remove pontos/traços
        cpf_limpo = limpar_cpf(cpf) 
        if not cpf_limpo:
             raise Exception("CPF inválido ou vazio.")

        servidor, dias_bruto, dias_desconto, dias_liquido, afastamentos = calcular_efetivo_exercicio(cpf_limpo)
        
        tempo_liquido_texto = converter_dias_para_texto(dias_liquido)
        
        return jsonify({
            "servidor": {
                "nome": servidor.nome,
                "cpf": servidor.cpf,
                "funcao": servidor.funcao or "Não informada",
                "data_admissao": servidor.data_inicio.strftime("%d/%m/%Y")
            },
            "tempo": {
                 "anos": dias_liquido // 365,
                 "meses": (dias_liquido % 365) // 30,
                 "dias": (dias_liquido % 365) % 30
            },
            "afastamentos": afastamentos
        })

    except Exception as e:
        # --- ATUALIZAÇÃO IMPORTANTE ---
        # Isso vai imprimir o erro exato no seu console (ex: "Servidor não encontrado...")
        print(f"ERRO 404 NA API da Calculadora: {e}") 
        # --- FIM DA ATUALIZAÇÃO ---
        
        return jsonify({"error": str(e)}), 404


@app.route("/certidao/efetivo_exercicio/<string:cpf>")
@login_required
@role_required("RH", "admin")
def gerar_certidao_efetivo_exercicio(cpf):
    """
    Rota que gera o PDF da Certidão.
    (VERSÃO v2.0 - Corrigida para receber 5 valores e gerar as tabelas)
    """
    try:
        cpf_limpo = limpar_cpf(cpf)
        
        # --- [ ESTA É A CORREÇÃO ] ---
        # Agora ela espera os 5 valores que a v4.0 envia:
        servidor, dias_bruto, dias_desconto, dias_liquido, afastamentos = calcular_efetivo_exercicio(cpf_limpo)
        # --- [ FIM DA CORREÇÃO ] ---

    except Exception as e:
        flash(f"Erro ao gerar certidão: {e}", "danger")
        return redirect(url_for("listar_requerimentos"))

    # 2. Prepara a geração do PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    story = []
    
    # --- ESTILOS DO PDF ---
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterBold', alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=14, spaceAfter=10))
    styles.add(ParagraphStyle(name='JustifyBody', alignment=TA_JUSTIFY, fontSize=11, leading=15, spaceAfter=12))
    styles.add(ParagraphStyle(name='CenterNormal', alignment=TA_CENTER, fontSize=10, leading=12))
    styles.add(ParagraphStyle(name='RightNormal', alignment=TA_RIGHT, fontSize=11, leading=14))
    styles.add(ParagraphStyle(name='TableHeader', alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=10, textColor=colors.black))
    styles.add(ParagraphStyle(name='TableCell', alignment=TA_LEFT, fontSize=9))
    styles.add(ParagraphStyle(name='TableCellCenter', alignment=TA_CENTER, fontSize=9))

    # --- INÍCIO DA CONSTRUÇÃO DO PDF ---
    
    story.append(Paragraph("CERTIDÃO DE EFETIVO EXERCÍCIO DO MAGISTÉRIO", styles['CenterBold']))
    story.append(Spacer(1, 0.5*cm))

    # --- Corpo do Texto ---
    texto_certidao = f"""
        CERTIFICAMOS, PARA FINS DE DIREITO que, de acordo com os elementos constantes do presente processo e,
        em especial, dos documentos e assentamentos de lavra da Secretaria Municipal de Educação que,
        <b>{servidor.nome.upper()}</b>,
        nascida em {servidor.data_nascimento.strftime('%d/%m/%Y') if servidor.data_nascimento else '[Data Nasc. não cadastrada]'},
        portador(a) do RG nº {servidor.rg or '[RG não cadastrado]'} SSP-PI,
        inscrita no CPF sob o nº {servidor.cpf},
        servidor(a) público(a) municipal, admitido(a) em <b>{servidor.data_inicio.strftime('%d/%m/%Y')}</b>,
        matrícula nº {servidor.num_contrato}, ocupante do cargo de <b>{servidor.funcao or '[Função não cadastrada]'}</b>,
        atualmente lotado(a) na {servidor.lotacao or '[Lotação não cadastrada]'}.
    """
    story.append(Paragraph(texto_certidao, styles['JustifyBody']))
    story.append(Spacer(1, 0.5*cm))

    # --- Tabela 1: Tempo de Efetivo Exercício (Bruto) ---
    data_tabela_bruto = [
        [Paragraph("Tempo de Efetivo Exercício no Magistério", styles['TableHeader']), Paragraph("Quantidade de Dias", styles['TableHeader'])],
        [
            Paragraph(f"Magistério: {servidor.data_inicio.strftime('%d/%m/%Y')} a {datetime.now().date().strftime('%d/%m/%Y')}", styles['TableCell']),
            [
                Paragraph(f"{dias_bruto} dias", styles['TableCellCenter']),
                Paragraph(converter_dias_para_texto(dias_bruto), styles['TableCellCenter']) # Usa a função auxiliar
            ]
        ],
        [
            Paragraph("Total", styles['TableHeader']),
            [
                Paragraph(f"{dias_bruto} dias", styles['TableCellCenter']),
                Paragraph(converter_dias_para_texto(dias_bruto), styles['TableCellCenter'])
            ]
        ]
    ]
    tabela_bruto = Table(data_tabela_bruto, colWidths=[10*cm, 7*cm])
    tabela_bruto.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E0E0E0")), # Cabeçalho cinza
        ('BACKGROUND', (0, 2), (0, 2), colors.HexColor("#E0E0E0")), # Total cinza
    ]))
    story.append(tabela_bruto)
    story.append(Spacer(1, 0.5*cm))

    # --- Tabela 2: Tempo Não Considerado (Descontos) ---
    data_tabela_descontos = [
        [Paragraph("Tempo não Considerado de Efetivo Exercício no Magistério", styles['TableHeader']), Paragraph("Quantidade de Dias", styles['TableHeader'])]
    ]
    
    if not afastamentos:
        data_tabela_descontos.append([Paragraph("Nenhum período de afastamento encontrado.", styles['TableCell']), Paragraph("0 dias", styles['TableCellCenter'])])
    
    for af in afastamentos:
        data_tabela_descontos.append([
            Paragraph(f"{af['data_inicio'].strftime('%d/%m/%Y')} a {af['data_fim'].strftime('%d/%m/%Y')} {af['natureza'].upper()}", styles['TableCell']),
            [
                Paragraph(f"{af['dias']} dias", styles['TableCellCenter']),
                Paragraph(converter_dias_para_texto(af['dias']), styles['TableCellCenter'])
            ]
        ])

    data_tabela_descontos.append([
        Paragraph("Total", styles['TableHeader']),
        [
            Paragraph(f"{dias_desconto} dias", styles['TableCellCenter']),
            Paragraph(converter_dias_para_texto(dias_desconto), styles['TableCellCenter'])
        ]
    ])
    
    tabela_descontos = Table(data_tabela_descontos, colWidths=[10*cm, 7*cm])
    tabela_descontos.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E0E0E0")), # Cabeçalho cinza
        ('BACKGROUND', (0, -1), (0, -1), colors.HexColor("#E0E0E0")), # Total cinza
    ]))
    story.append(tabela_descontos)
    story.append(Spacer(1, 0.5*cm))

    # --- Tabela 3: Resumo (Líquido) ---
    data_tabela_resumo = [
        [
            Paragraph("Total Bruto", styles['TableCell']), 
            Paragraph(f"{dias_bruto} dias", styles['TableCellCenter']), 
            Paragraph(converter_dias_para_texto(dias_bruto), styles['TableCellCenter'])
        ],
        [
            Paragraph("Desconto", styles['TableCell']),
            Paragraph(f"{dias_desconto} dias", styles['TableCellCenter']),
            Paragraph(converter_dias_para_texto(dias_desconto), styles['TableCellCenter'])
        ],
        [
            Paragraph("Líquido de Efetivo Exercício do Magistério", styles['TableHeader']),
            Paragraph(f"{dias_liquido} dias", styles['TableCellCenter']),
            Paragraph(converter_dias_para_texto(dias_liquido), styles['TableCellCenter'])
        ]
    ]
    tabela_resumo = Table(data_tabela_resumo, colWidths=[7*cm, 3*cm, 7*cm])
    tabela_resumo.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 2), (0, 2), colors.HexColor("#E0E0E0")), # Linha do líquido cinza
    ]))
    story.append(tabela_resumo)
    story.append(Spacer(1, 1*cm))

    # --- Data e Assinatura ---
    data_hoje_extenso = datetime.now().strftime("%d de %B de %Y")
    story.append(Paragraph(f"Valença do Piauí, {data_hoje_extenso}", styles['RightNormal']))
    story.append(Spacer(1, 2*cm))
    
    story.append(HRFlowable(width="70%", thickness=0.5, color=colors.black, hAlign="CENTER"))
    story.append(Paragraph("Assinatura do Responsável", styles['CenterNormal']))
    
    # --- Fim da Construção ---
    doc.build(story)
    buffer.seek(0)
    
    registrar_log(f'Gerou Certidão de Efetivo Exercício (Layout .doc) para {servidor.nome}.')

    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=Certidao_Efetivo_Exercicio_{servidor.nome.replace(" ", "_")}.pdf'
    
    return response

@app.route("/combustivel", methods=["GET", "POST"])
@login_required
@role_required("Combustivel", "admin")
def lancar_abastecimento():
    if request.method == "POST":
        try:
            litros_str = request.form.get("litros", "").strip().replace(",", ".")
            valor_litro_str = (
                request.form.get("valor_litro", "").strip().replace(",", ".")
            )
            quilometragem_str = (
                request.form.get("quilometragem", "").strip().replace(",", ".")
            )
            litros = float(litros_str) if litros_str else 0.0
            valor_litro = float(valor_litro_str) if valor_litro_str else 0.0
            quilometragem_val = int(quilometragem_str) if quilometragem_str else 0
            valor_total = litros * valor_litro

            novo_abastecimento = Abastecimento(
                veiculo_placa=request.form.get("veiculo_placa"),
                motorista_id=request.form.get("motorista_id"),
                quilometragem=quilometragem_val,
                tipo_combustivel=request.form.get("tipo_combustivel"),
                litros=litros,
                valor_litro=valor_litro,
                valor_total=valor_total,
            )
            db.session.add(novo_abastecimento)
            db.session.commit()
            flash("Abastecimento registrado com sucesso!", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao registrar abastecimento: {e}", "danger")

        return redirect(url_for("lancar_abastecimento"))

    veiculos = Veiculo.query.order_by(Veiculo.modelo).all()
    motoristas = Motorista.query.order_by(Motorista.nome).all()
    abastecimentos = (
        Abastecimento.query.order_by(Abastecimento.data.desc()).limit(15).all()
    )

    return render_template(
        "combustivel.html",  # ATENÇÃO: Verifique se este é o nome correto do seu template
        veiculos=veiculos,
        motoristas=motoristas,
        abastecimentos=abastecimentos,
    )




# Adicione também as outras rotas relacionadas se estiverem faltando, como a de registrar o ponto
@app.route("/ponto/registrar", methods=["GET", "POST"])
def registrar_ponto():
    if request.method == "POST":
        try:
            foto_b64 = request.form.get("foto_b64")
            tipo_registro = request.form.get("tipo")
            escola_id_form = request.form.get("escola_id", type=int) # Escola onde ele ESTÁ tentando bater
            lat_user_str = request.form.get("latitude")
            lon_user_str = request.form.get("longitude")

            if not foto_b64 or not escola_id_form:
                flash("Foto ou local de trabalho não detectados.", "danger")
                return redirect(url_for("registrar_ponto"))

            # 1. Identificação Facial
            todos_servidores = Servidor.query.filter(Servidor.face_encoding.isnot(None)).all()
            servidor_identificado, msg_identificacao = identificar_servidor_por_rosto(foto_b64, todos_servidores)

            if not servidor_identificado:
                flash(f"Falha na identificação: {msg_identificacao}", "danger")
                return redirect(url_for("registrar_ponto"))

            # --- NOVO BLOQUEIO: VÍNCULO DE ESCOLA ---
            # Se o servidor tem uma escola vinculada E a escola selecionada é diferente
            if servidor_identificado.escola_id and servidor_identificado.escola_id != escola_id_form:
                escola_correta = servidor_identificado.escola_vinculada.nome if servidor_identificado.escola_vinculada else "outra unidade"
                flash(f"ACESSO NEGADO: Você está lotado na '{escola_correta}'. Não é permitido registrar ponto nesta localização.", "danger")
                return redirect(url_for("registrar_ponto"))
            # ----------------------------------------

            # 2. Validação de Geolocalização (GPS)
            escola_local = Escola.query.get(escola_id_form)
            
            if lat_user_str == 'N/A' or lon_user_str == 'N/A':
                flash("Erro: GPS não detectado. Ative a localização.", "warning")
                return redirect(url_for("registrar_ponto"))

            if escola_local and escola_local.latitude:
                try:
                    lat_user = float(lat_user_str)
                    lon_user = float(lon_user_str)
                    distancia = haversine(lat_user, lon_user, escola_local.latitude, escola_local.longitude)
                    limite_metros = app.config.get('RAIO_PERMITIDO_METROS', 100)

                    if distancia > limite_metros:
                        flash(f"Fora do perímetro! Distância: {distancia:.0f}m. Limite: {limite_metros}m.", "danger")
                        return redirect(url_for("registrar_ponto"))

                except Exception as e:
                    print(f"Erro GPS: {e}")
                    flash("Erro ao validar localização.", "danger")
                    return redirect(url_for("registrar_ponto"))

            # 3. Salvar o Ponto
            filename_ponto = f"ponto_{servidor_identificado.num_contrato}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
            
            novo_ponto = Ponto(
                servidor_cpf=servidor_identificado.cpf,
                tipo=tipo_registro,
                escola_id=escola_id_form,
                latitude=float(lat_user_str),
                longitude=float(lon_user_str),
                foto_filename=filename_ponto
            )
            
            db.session.add(novo_ponto)
            db.session.commit()

            hora = datetime.now().strftime('%H:%M')
            nome = servidor_identificado.nome.split()[0]
            flash(f"Sucesso, {nome}! {tipo_registro.capitalize()} registrada às {hora}.", "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"Erro no sistema: {e}", "danger")

        return redirect(url_for("registrar_ponto"))

    escolas = Escola.query.filter_by(status="Ativa").order_by(Escola.nome).all()
    return render_template("registrar_ponto.html", escolas=escolas)

@app.route("/bloco_de_notas")
@login_required
@role_required("RH", "admin")
def bloco_de_notas():
    user = User.query.filter_by(username=session["username"]).first_or_404()
    notas = (
        Nota.query.filter_by(user_id=user.id).order_by(Nota.data_criacao.desc()).all()
    )
    return render_template("bloco_de_notas.html", notas=notas)


@app.route("/notas/add", methods=["POST"])
@login_required
def add_nota():
    user = User.query.filter_by(username=session["username"]).first_or_404()
    titulo = request.form.get("titulo")
    conteudo = request.form.get("conteudo")
    if not titulo:
        flash("O título da anotação é obrigatório.", "warning")
        return redirect(url_for("bloco_de_notas"))
    nova_nota = Nota(titulo=titulo, conteudo=conteudo, autor=user)
    db.session.add(nova_nota)
    db.session.commit()
    registrar_log(f'Criou a anotação: "{titulo}"')
    flash("Anotação criada com sucesso!", "success")
    return redirect(url_for("bloco_de_notas"))


@app.route("/notas/update/<int:id>", methods=["POST"])
@login_required
def update_nota(id):
    nota = Nota.query.get_or_404(id)
    if nota.autor.username != session["username"]:
        abort(403)
    nota.titulo = request.form.get("titulo")
    nota.conteudo = request.form.get("conteudo")
    db.session.commit()
    registrar_log(f'Editou a anotação: "{nota.titulo}"')
    flash("Anotação atualizada com sucesso!", "success")
    return redirect(url_for("bloco_de_notas"))


@app.route("/notas/delete/<int:id>")
@login_required
def delete_nota(id):
    nota = Nota.query.get_or_404(id)
    if nota.autor.username != session["username"]:
        abort(403)
    titulo_nota = nota.titulo
    db.session.delete(nota)
    db.session.commit()
    registrar_log(f'Excluiu a anotação: "{titulo_nota}"')
    flash("Anotação excluída com sucesso!", "success")
    return redirect(url_for("bloco_de_notas"))


@app.route("/combustivel/relatorio", methods=["GET"])
@login_required
@role_required("Combustivel", "admin")
def relatorio_combustivel():
    # Coleta os filtros da URL
    placa_filtro = request.args.get("placa")
    query = (
        db.session.query(Abastecimento)
        .join(Veiculo)
        .order_by(Abastecimento.veiculo_placa, Abastecimento.quilometragem)
    )

    if placa_filtro:
        query = query.filter(Veiculo.placa == placa_filtro)

    resultados_filtrados = query.all()
    resultados_com_analise = []
    ultimo_abastecimento = {}

    for r in resultados_filtrados:
        placa = r.veiculo_placa
        analise = {"abastecimento": r, "km_rodado": 0, "consumo_kml": 0, "custo_km": 0}

        if placa in ultimo_abastecimento:
            anterior = ultimo_abastecimento[placa]
            km_rodado = r.quilometragem - anterior.quilometragem
            analise["km_rodado"] = km_rodado
            if km_rodado > 0 and anterior.litros > 0:
                analise["consumo_kml"] = km_rodado / anterior.litros
                analise["custo_km"] = anterior.valor_total / km_rodado

        resultados_com_analise.append(analise)
        ultimo_abastecimento[placa] = r

    veiculos = Veiculo.query.order_by(Veiculo.modelo).all()

    return render_template(
        "relatorio_combustivel.html",
        resultados=resultados_com_analise,
        veiculos=veiculos,
    )


@app.route("/combustivel/excluir/<int:id>")
@login_required
def excluir_abastecimento(id):
    abastecimento_para_excluir = Abastecimento.query.get_or_404(id)
    try:
        info_log = f"veículo placa {abastecimento_para_excluir.veiculo_placa}, {abastecimento_para_excluir.litros}L em {abastecimento_para_excluir.data.strftime('%d/%m/%Y')}"
        db.session.delete(abastecimento_para_excluir)
        db.session.commit()
        registrar_log(f"Excluiu o registro de abastecimento: {info_log}.")
        flash("Registro de abastecimento excluído com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir o registro: {e}", "danger")

    return redirect(url_for("lancar_abastecimento"))


@app.route("/ponto/qrcode")
@login_required
@admin_required
def exibir_qrcode_ponto():
    url_registro = url_for("registrar_ponto", _external=True)
    img = qrcode.make(url_registro)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    img_str = base64.b64encode(buf.getvalue()).decode("utf-8")
    registrar_log("Gerou o QR Code para registro de ponto.")
    return render_template("qrcode_ponto.html", qr_code_image=img_str)


@app.route("/debug-sessao")
def debug_sessao():
    return dict(session)


@app.route("/debug/cpf/<string:cpf_busca>")
@login_required
def debug_cpf(cpf_busca):
    try:
        # Limpa o CPF da URL (como o JS faz)
        cpf_limpo = re.sub(r"\D", "", cpf_busca)
        
        print(f"--- INICIANDO DEBUG DE CPF PARA: {cpf_limpo} ---")
        
        # Tenta a nossa query "inteligente" (v3.2)
        servidor_encontrado = Servidor.query.filter(
            func.replace(func.replace(func.replace(Servidor.cpf, '.', ''), '-', ''), ' ', '') == cpf_limpo
        ).first()

        if servidor_encontrado:
            # Se encontrou, ótimo!
            print(f"DEBUG (Debug Route): SUCESSO. Encontrou: {servidor_encontrado.nome}")
            print(f"DEBUG (Debug Route): CPF no DB (entre aspas): '{servidor_encontrado.cpf}'")
            return f"<h1>SUCESSO!</h1><p>Encontrou: {servidor_encontrado.nome}</p><p>CPF no DB (veja as aspas): '{servidor_encontrado.cpf}'</p>"
        
        # Se não encontrou, vamos tentar uma busca "suja" (LIKE)
        print(f"DEBUG (Debug Route): FALHA na query v3.2. Tentando busca 'LIKE'...")
        servidor_like = Servidor.query.filter(Servidor.cpf.ilike(f"%{cpf_limpo}%")).all()
        
        if servidor_like:
            resultados = f"<h1>FALHA NA QUERY v3.2</h1><p>Mas uma busca 'LIKE' encontrou (veja os CPFs no banco):</p>"
            for s in servidor_like:
                # Mostra o CPF exatamente como está, com ' quotes para ver os espaços
                resultados += f"<li>Nome: {s.nome}, CPF no DB: '{s.cpf}'</li>"
                print(f"DEBUG (Debug Route): Encontrado com LIKE: Nome: {s.nome}, CPF no DB: '{s.cpf}'")
            return resultados
        
        print(f"DEBUG (Debug Route): FALHA TOTAL. Nenhum servidor encontrado.")
        return f"<h1>FALHA TOTAL</h1><p>Nenhum servidor encontrado com o CPF '{cpf_limpo}' ou algo parecido no banco de dados.</p>"
        
    except Exception as e:
        return f"<h1>ERRO NO DEBUG</h1><p>{str(e)}</p>"
    

@app.route("/ponto/frequencia")
@login_required
@admin_required
@role_required("RH", "admin")
def visualizar_frequencia():
    # --- LÓGICA DE FILTRO ---
    page = request.args.get("page", 1, type=int)
    
    # Pega os filtros da URL (do formulário GET)
    filtro_mes = request.args.get("mes", type=int)
    filtro_ano = request.args.get("ano", type=int)
    filtro_escola_id = request.args.get("escola_id", type=int)

    # Começa a query base
    query = Ponto.query.join(Servidor, Ponto.servidor_cpf == Servidor.cpf)\
                       .join(Escola, Ponto.escola_id == Escola.id)\
                       .order_by(Ponto.timestamp.desc())

    # Aplica os filtros se eles existirem
    if filtro_mes:
        query = query.filter(db.extract('month', Ponto.timestamp) == filtro_mes)
    if filtro_ano:
        query = query.filter(db.extract('year', Ponto.timestamp) == filtro_ano)
    if filtro_escola_id:
        query = query.filter(Ponto.escola_id == filtro_escola_id)

    # Executa a query com paginação
    registros_paginados = query.paginate(page=page, per_page=50, error_out=False)
    
    # --- DADOS PARA OS FILTROS ---
    # Busca todas as escolas para popular o dropdown
    escolas = Escola.query.order_by(Escola.nome).all()
    
    # Gera uma lista de anos (do ano atual até 5 anos atrás)
    ano_atual = datetime.now().year
    anos_disponiveis = list(range(ano_atual, ano_atual - 6, -1))

    return render_template(
        "frequencia.html",
        registros=registros_paginados,
        escolas=escolas,
        anos_disponiveis=anos_disponiveis,
        # Passa os filtros atuais de volta para o template
        filtros_atuais={
            'mes': filtro_mes,
            'ano': filtro_ano,
            'escola_id': filtro_escola_id
        }
    )

def _get_dados_frequencia_filtrados():
    """Função auxiliar para evitar código duplicado nas rotas de exportação."""
    
    filtro_mes = request.args.get("mes", type=int)
    filtro_ano = request.args.get("ano", type=int)
    filtro_escola_id = request.args.get("escola_id", type=int)

    query = Ponto.query.join(Servidor, Ponto.servidor_cpf == Servidor.cpf)\
                       .join(Escola, Ponto.escola_id == Escola.id)\
                       .order_by(Ponto.timestamp.asc()) # ASC para relatórios cronológicos

    if filtro_mes:
        query = query.filter(db.extract('month', Ponto.timestamp) == filtro_mes)
    if filtro_ano:
        query = query.filter(db.extract('year', Ponto.timestamp) == filtro_ano)
    if filtro_escola_id:
        query = query.filter(Ponto.escola_id == filtro_escola_id)
        
    return query.all()

@app.route("/ponto/exportar/excel")
@login_required
@admin_required
@role_required("RH", "admin")
def exportar_frequencia_excel():
    registros = _get_dados_frequencia_filtrados()

    wb = Workbook()
    ws = wb.active
    ws.title = "Frequência"

    # --- Cabeçalho ---
    headers = ["Servidor", "CPF", "Escola", "Data", "Hora", "Tipo"]
    ws.append(headers)
    for cell in ws[1]: # Itera sobre a primeira linha (cabeçalho)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # --- Dados ---
    for reg in registros:
        ws.append([
            reg.servidor_ponto.nome if reg.servidor_ponto else "N/A",
            reg.servidor_cpf,
            reg.escola.nome if reg.escola else "N/A",
            reg.timestamp.strftime('%d/%m/%Y'),
            reg.timestamp.strftime('%H:%M:%S'),
            reg.tipo.capitalize()
        ])
        
    # Ajusta a largura das colunas
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].autosize = True

    # --- Salva e Envia ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=relatorio_frequencia_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    registrar_log("Gerou relatório de frequência em Excel.")
    return response


@app.route("/ponto/exportar/pdf")
@login_required
@admin_required
@role_required("RH", "admin")
def exportar_frequencia_pdf():
    registros = _get_dados_frequencia_filtrados()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=2.5*cm, bottomMargin=2.5*cm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Relatório de Frequência", styles['h1']),
        Spacer(1, 0.5*cm)
    ]

    # --- Cabeçalho da Tabela ---
    header_style = ParagraphStyle(name='Header', fontSize=9, fontName='Helvetica-Bold', alignment=TA_CENTER)
    data = [[
        Paragraph("Servidor", header_style),
        Paragraph("CPF", header_style),
        Paragraph("Escola", header_style),
        Paragraph("Data", header_style),
        Paragraph("Hora", header_style),
        Paragraph("Tipo", header_style)
    ]]

    # --- Dados da Tabela ---
    cell_style = ParagraphStyle(name='Cell', fontSize=8, alignment=TA_CENTER)
    for reg in registros:
        data.append([
            Paragraph(reg.servidor_ponto.nome if reg.servidor_ponto else "N/A", cell_style),
            Paragraph(reg.servidor_cpf, cell_style),
            Paragraph(reg.escola.nome if reg.escola else "N/A", cell_style),
            Paragraph(reg.timestamp.strftime('%d/%m/%Y'), cell_style),
            Paragraph(reg.timestamp.strftime('%H:%M:%S'), cell_style),
            Paragraph(reg.tipo.capitalize(), cell_style)
        ])
    
    table = Table(data, colWidths=[8*cm, 3*cm, 7*cm, 2.5*cm, 2.5*cm, 2*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#004d40")), # Cor do seu cabeçalho
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(table)
    
    # Adicionamos o cabeçalho e rodapé padrão
    doc.build(story, onFirstPage=cabecalho_e_rodape, onLaterPages=cabecalho_e_rodape)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers["Content-Disposition"] = f"inline; filename=relatorio_frequencia_{datetime.now().strftime('%Y%m%d')}.pdf"
    response.headers["Content-Type"] = "application/pdf"
    
    registrar_log("Gerou relatório de frequência em PDF.")
    return response


def _get_dados_frequencia_individual(cpf, mes=None, ano=None):
    """
    Função auxiliar que busca os dados de ponto para um CPF específico,
    opcionalmente filtrando por mês e ano.
    """
    # 1. Encontra o servidor (precisamos do nome dele para o relatório)
    #    (Assumindo que o CPF é único, 'first()' é seguro)
    servidor = Servidor.query.filter_by(cpf=cpf).first()
    
    if not servidor:
        # Se o CPF não corresponder a nenhum servidor
        return None, None 

    # 2. Constrói a query de Ponto
    query = Ponto.query.filter_by(servidor_cpf=cpf)\
                       .join(Escola, Ponto.escola_id == Escola.id, isouter=True)\
                       .order_by(Ponto.timestamp.asc()) # Ordena do mais antigo para o mais novo

    # 3. Aplica filtros de data, se existirem
    if mes:
        query = query.filter(db.extract('month', Ponto.timestamp) == mes)
    if ano:
        query = query.filter(db.extract('year', Ponto.timestamp) == ano)
    
    # 4. Retorna os registros encontrados E o objeto do servidor
    return query.all(), servidor

@app.route("/servidor/<string:cpf>/exportar/excel")
@login_required
@admin_required
@role_required("RH", "admin")
def exportar_frequencia_individual_excel(cpf):
    # Pega os filtros opcionais de Mês e Ano da URL
    filtro_mes = request.args.get("mes", type=int)
    filtro_ano = request.args.get("ano", type=int)
    
    registros, servidor = _get_dados_frequencia_individual(cpf, filtro_mes, filtro_ano)

    if not servidor:
        flash(f"Servidor com CPF {cpf} não encontrado.", "danger")
        return redirect(url_for("lista_servidores"))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Frequência - {servidor.nome[:30]}" # Limita o nome da aba

    # --- Título do Relatório ---
    ws.merge_cells('A1:E1')
    titulo_cell = ws['A1']
    titulo_cell.value = f"Relatório de Frequência Individual - {servidor.nome}"
    titulo_cell.font = Font(bold=True, size=16)
    titulo_cell.alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:E2')
    subtitulo_cell = ws['A2']
    subtitulo_cell.value = f"CPF: {servidor.cpf} | Mês/Ano: {filtro_mes or 'Todos'}/{filtro_ano or 'Todos'}"
    subtitulo_cell.font = Font(italic=True, size=12)
    subtitulo_cell.alignment = Alignment(horizontal="center")
    ws.append([]) # Linha em branco

    # --- Cabeçalho da Tabela ---
    headers = ["Local (Escola)", "Data", "Hora", "Tipo", "Auditoria"]
    ws.append(headers)
    for cell in ws[4]: # Cabeçalho está na linha 4
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # --- Dados ---
    for reg in registros:
        # Verifica se a auditoria de foto existe
        auditoria_foto = "Sim" if (reg.servidor_ponto.foto_filename and reg.foto_filename) else "Não"
        
        ws.append([
            reg.escola.nome if reg.escola else "N/A",
            reg.timestamp.strftime('%d/%m/%Y'),
            reg.timestamp.strftime('%H:%M:%S'),
            reg.tipo.capitalize(),
            auditoria_foto
        ])
        
    # Ajusta a largura das colunas
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].autosize = True

    # --- Salva e Envia ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    filename = f"frequencia_{servidor.nome.replace(' ', '_')}.xlsx"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    registrar_log(f"Gerou relatório Excel de frequência para {servidor.nome}.")
    return response


@app.route("/servidor/<string:cpf>/exportar/pdf")
@login_required
@admin_required
@role_required("RH", "admin")
def exportar_frequencia_individual_pdf(cpf):
    filtro_mes = request.args.get("mes", type=int)
    filtro_ano = request.args.get("ano", type=int)

    registros, servidor = _get_dados_frequencia_individual(cpf, filtro_mes, filtro_ano)

    if not servidor:
        flash(f"Servidor com CPF {cpf} não encontrado.", "danger")
        return redirect(url_for("lista_servidores"))

    buffer = io.BytesIO()
    # Usamos A4 normal (retrato), não landscape
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2.5*cm, bottomMargin=2.5*cm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Relatório de Frequência Individual", styles['h1']),
        Paragraph(f"<b>Servidor:</b> {servidor.nome}", styles['h2']),
        Paragraph(f"<b>CPF:</b> {servidor.cpf}", styles['Normal']),
        Paragraph(f"<b>Período:</b> {filtro_mes or 'Todos os meses'}/{filtro_ano or 'Todos os anos'}", styles['Normal']),
        Spacer(1, 1*cm)
    ]

    # --- Cabeçalho da Tabela ---
    header_style = ParagraphStyle(name='Header', fontSize=9, fontName='Helvetica-Bold', alignment=TA_CENTER)
    data = [[
        Paragraph("Local (Escola)", header_style),
        Paragraph("Data", header_style),
        Paragraph("Hora", header_style),
        Paragraph("Tipo", header_style)
    ]]

    # --- Dados da Tabela ---
    cell_style = ParagraphStyle(name='Cell', fontSize=8, alignment=TA_CENTER)
    for reg in registros:
        data.append([
            Paragraph(reg.escola.nome if reg.escola else "N/A", cell_style),
            Paragraph(reg.timestamp.strftime('%d/%m/%Y'), cell_style),
            Paragraph(reg.timestamp.strftime('%H:%M:%S'), cell_style),
            Paragraph(reg.tipo.capitalize(), cell_style)
        ])
    
    if not registros:
        story.append(Paragraph("Nenhum registro de ponto encontrado para este período.", styles['Normal']))
    else:
        table = Table(data, colWidths=[8*cm, 3*cm, 3*cm, 3*cm]) # Ajustado para A4 retrato
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#004d40")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(table)
    
    # Adicionamos o cabeçalho e rodapé padrão
    doc.build(story, onFirstPage=cabecalho_e_rodape, onLaterPages=cabecalho_e_rodape)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    filename = f"frequencia_{servidor.nome.replace(' ', '_')}.pdf"
    response.headers["Content-Disposition"] = f"inline; filename={filename}"
    response.headers["Content-Type"] = "application/pdf"
    
    registrar_log(f"Gerou relatório PDF de frequência para {servidor.nome}.")
    return response

def haversine(lat1, lon1, lat2, lon2):
    """
    Calcula a distância do grande círculo entre dois pontos na Terra 
    (especificados em graus decimais). Retorna a distância em metros.
    """
    R = 6371000  # Raio da Terra em metros
    
    # Converte graus para radianos
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)
    
    # Fórmula Haversine
    dlon = lon2_rad - lon1_rad
    dlat = lat2_rad - lat1_rad
    
    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    
    distance = R * c
    return distance

# --- ROTA TEMPORÁRIA PARA CRIAR TABELAS QUE FALTAM ---
@app.route('/criar-tabelas-extras')
def criar_tabelas_extras():
    try:
        from models import RelatorioTecnico, RelatorioAnexo
        with app.app_context():
            db.create_all()
            return """
            <div style='text-align: center; padding: 50px; font-family: Arial;'>
                <h1 style='color: green;'>Sucesso! ✅</h1>
                <p>A tabela <b>relatorios_tecnicos</b> foi criada no banco de dados.</p>
                <a href='/merenda/dashboard' style='padding: 10px 20px; background: #007bff; color: white; text-decoration: none; border-radius: 5px;'>Voltar para o Sistema</a>
            </div>
            """
    except Exception as e:
        return f"<h1>Erro ao criar tabelas: {e}</h1>"

# ===================================================================
# PARTE 6: Importação e Registro dos Blueprints
# ===================================================================
from patrimonio_routes import patrimonio_bp
from merenda_routes import merenda_bp
from motoristas_routes import motoristas_bp
from escola_routes import escola_bp
from transporte_routes import transporte_bp
from protocolo_routes import protocolo_bp
from contratos_routes import contratos_bp
from frequencia_routes import frequencia_bp
from backup_routes import backup_bp
from almoxarifado_routes import almoxarifado_bp
from academico_routes import academico_bp
from caee_routes import caee_bp 
from contrato_fiscal_routes import contrato_fiscal_bp
from contas_routes import contas_bp
from whatsapp_routes import whatsapp_bp
from assinatura_routes import assinatura_bp


app.register_blueprint(transporte_bp)
app.register_blueprint(protocolo_bp)
app.register_blueprint(contratos_bp)
app.register_blueprint(patrimonio_bp)
app.register_blueprint(merenda_bp)
app.register_blueprint(motoristas_bp)
app.register_blueprint(escola_bp)
app.register_blueprint(frequencia_bp)
app.register_blueprint(backup_bp)
app.register_blueprint(almoxarifado_bp)
app.register_blueprint(academico_bp)
app.register_blueprint(caee_bp)
app.register_blueprint(contrato_fiscal_bp)
app.register_blueprint(contas_bp)
app.register_blueprint(whatsapp_bp)
app.register_blueprint(assinatura_bp)

# ===================================================================
# PARTE 7: Bloco de Execução Principal
# ===================================================================


if __name__ == "__main__":
    app.run(debug=True)
